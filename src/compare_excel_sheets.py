import pandas as pd
import argparse
import ast
import logging
import math
import operator
import re
from decimal import Decimal, InvalidOperation
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import column_index_from_string

# Setup logging configuration
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%H:%M:%S'
)

MISMATCH_OUTPUT_THRESHOLD = 5
CELL_REF_PATTERN = re.compile(r"(?<![A-Z0-9_])\$?([A-Z]{1,3})\$?([0-9]+)\b")
FORMULA_CELL_REF_PATTERN = re.compile(r"(?<![A-Z0-9_])(\$?)([A-Z]{1,3})(\$?)([0-9]+)\b")
FULL_CELL_REF_PATTERN = re.compile(r"^\$?([A-Z]{1,3})\$?([0-9]+)$")
ARITHMETIC_EXPRESSION_PATTERN = re.compile(r"^[0-9+\-*/().\s]+$")
ALLOWED_BINARY_OPERATORS = {
    ast.Add: operator.add,
    ast.Sub: operator.sub,
    ast.Mult: operator.mul,
    ast.Div: operator.truediv,
}
ALLOWED_UNARY_OPERATORS = {
    ast.UAdd: operator.pos,
    ast.USub: operator.neg,
}
INTEGER_STRING_PATTERN = re.compile(r"^[+-]?\d+$")


def normalize_decimal_string(value):
    if value.is_nan():
        return ""

    if not value.is_finite():
        return str(value)

    decimal_text = format(value, "f")
    if "." not in decimal_text:
        return decimal_text

    return decimal_text.rstrip("0").rstrip(".")


def normalize_scalar(value):
    if value is None:
        return ""

    if isinstance(value, str):
        value = value.strip()
        if value.lower() == "nan":
            return ""
        if INTEGER_STRING_PATTERN.match(value):
            return value

        try:
            numeric = Decimal(value)
        except InvalidOperation:
            return value

        return normalize_decimal_string(numeric)

    if pd.isna(value):
        return ""

    if isinstance(value, bool):
        return str(value)

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float):
        if math.isnan(value):
            return ""
        return normalize_decimal_string(Decimal(str(value)))

    return str(value).strip()


def safe_eval_arithmetic(expression):
    def eval_node(node):
        if isinstance(node, ast.Expression):
            return eval_node(node.body)
        if isinstance(node, ast.Constant) and isinstance(node.value, (int, float)):
            return float(node.value)
        if isinstance(node, ast.BinOp) and type(node.op) in ALLOWED_BINARY_OPERATORS:
            left = eval_node(node.left)
            right = eval_node(node.right)
            return ALLOWED_BINARY_OPERATORS[type(node.op)](left, right)
        if isinstance(node, ast.UnaryOp) and type(node.op) in ALLOWED_UNARY_OPERATORS:
            return ALLOWED_UNARY_OPERATORS[type(node.op)](eval_node(node.operand))
        raise ValueError("Unsupported formula expression")

    parsed = ast.parse(expression, mode="eval")
    return eval_node(parsed)


def load_excel_sheet_data(path, sheet_name):
    workbook_values = load_workbook(path, data_only=True, read_only=True)
    workbook_with_formulas = load_workbook(path, data_only=False, read_only=True)
    ws_values = workbook_values[sheet_name]
    ws_formulas = workbook_with_formulas[sheet_name]
    resolved_headers = resolve_headers_from_worksheet(ws_formulas)

    data_rows = [list(row) for row in ws_values.iter_rows(min_row=2, values_only=True)]
    current_column_count = max((len(row) for row in data_rows), default=0)
    df = pd.DataFrame(data_rows)

    if current_column_count > len(resolved_headers):
        resolved_headers.extend(
            get_excel_column_label(idx + 1)
            for idx in range(len(resolved_headers), current_column_count)
        )

    expected_column_count = len(resolved_headers)
    if current_column_count < expected_column_count:
        for idx in range(current_column_count, expected_column_count):
            df[idx] = pd.NA

    df.columns = resolved_headers[:len(df.columns)]
    formula_df = pd.DataFrame("", index=df.index, columns=df.columns)

    for row_idx, row in enumerate(
        ws_formulas.iter_rows(min_row=2, max_row=len(df) + 1),
        start=0,
    ):
        for col_idx, cell in enumerate(row, start=0):
            if col_idx >= len(df.columns):
                break
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_df.iat[row_idx, col_idx] = normalize_formula(cell.value, row_idx + 2)
            if isinstance(cell.value, str) and not cell.value.startswith("="):
                df.iat[row_idx, col_idx] = cell.value
                continue
            if not pd.isna(df.iat[row_idx, col_idx]):
                continue
            if not isinstance(cell.value, str) or not cell.value.startswith("="):
                continue

            resolved_value = resolve_formula_from_dataframe(cell.value, df, row_idx)
            if resolved_value is not None:
                df.iat[row_idx, col_idx] = resolved_value

    workbook_values.close()
    workbook_with_formulas.close()
    return df, formula_df


def load_excel_dataframe(path, sheet_name):
    df, _ = load_excel_sheet_data(path, sheet_name)
    return df


def load_physical_column_values(path, sheet_name, column_letter, row_count):
    workbook_values = load_workbook(path, data_only=True, read_only=True)
    workbook_with_formulas = load_workbook(path, data_only=False, read_only=True)
    ws_values = workbook_values[sheet_name]
    ws_formulas = workbook_with_formulas[sheet_name]
    values = []

    for row_number in range(2, row_count + 2):
        formula_cell = ws_formulas[f"{column_letter}{row_number}"]
        value_cell = ws_values[f"{column_letter}{row_number}"]

        if isinstance(formula_cell.value, str) and not formula_cell.value.startswith("="):
            raw_value = formula_cell.value
        else:
            raw_value = value_cell.value

        values.append(raw_value)

    workbook_values.close()
    workbook_with_formulas.close()
    return values


def get_excel_column_label(column_index):
    label = ""
    while column_index:
        column_index, remainder = divmod(column_index - 1, 26)
        label = chr(65 + remainder) + label
    return label


def normalize_formula(formula, row_number):
    if not isinstance(formula, str) or not formula.startswith("="):
        return ""

    def replace_cell_ref(match):
        col_abs, col_label, row_abs, ref_row = match.groups()
        if row_abs:
            normalized_row = f"${ref_row}"
        else:
            normalized_row = f"[{int(ref_row) - row_number}]"

        return f"{col_abs}{col_label}{normalized_row}"

    formula = FORMULA_CELL_REF_PATTERN.sub(replace_cell_ref, formula.strip())
    return remove_formula_whitespace(formula)


def remove_formula_whitespace(formula):
    result = []
    in_string = False

    for char in formula:
        if char == '"':
            in_string = not in_string
            result.append(char)
            continue
        if in_string or not char.isspace():
            result.append(char)

    return "".join(result)


def make_unique_headers(headers):
    seen = {}
    unique_headers = []

    for header in headers:
        header = str(header).strip()
        count = seen.get(header, 0)
        if count:
            unique_headers.append(f"{header}.{count}")
        else:
            unique_headers.append(header)
        seen[header] = count + 1

    return unique_headers


def resolve_headers_from_worksheet(ws_formulas):
    raw_headers = [cell.value for cell in next(ws_formulas.iter_rows(min_row=1, max_row=1))]
    cache = {}

    def resolve_header(idx, stack):
        if idx in cache:
            return cache[idx]
        if idx in stack:
            return ""

        value = raw_headers[idx]
        if value is None:
            resolved = ""
        elif isinstance(value, str) and value.startswith("="):
            resolved = resolve_header_formula(value, raw_headers, stack | {idx})
        else:
            resolved = str(value).strip()

        cache[idx] = resolved
        return resolved

    resolved_headers = []
    for idx in range(len(raw_headers)):
        resolved = resolve_header(idx, set())
        if resolved == "":
            continue
        resolved_headers.append(resolved)

    return make_unique_headers(resolved_headers)


def resolve_header_formula(formula, raw_headers, stack):
    expression = formula[1:].strip()

    direct_ref_match = FULL_CELL_REF_PATTERN.fullmatch(expression)
    if direct_ref_match:
        ref_col = column_index_from_string(direct_ref_match.group(1)) - 1
        ref_row = int(direct_ref_match.group(2))
        if ref_row != 1:
            raise ValueError("Header formulas may only reference row 1")
        return resolve_header_operand(ref_col, raw_headers, stack)

    if "&" in expression:
        parts = [part.strip() for part in expression.split("&")]
        return "".join(resolve_header_formula_part(part, raw_headers, stack) for part in parts)

    raise ValueError("Unsupported header formula")


def resolve_header_formula_part(part, raw_headers, stack):
    ref_match = FULL_CELL_REF_PATTERN.fullmatch(part)
    if ref_match:
        ref_col = column_index_from_string(ref_match.group(1)) - 1
        ref_row = int(ref_match.group(2))
        if ref_row != 1:
            raise ValueError("Header formulas may only reference row 1")
        return resolve_header_operand(ref_col, raw_headers, stack)

    if part.startswith('"') and part.endswith('"'):
        return part[1:-1]

    raise ValueError("Unsupported header formula part")


def resolve_header_operand(ref_col, raw_headers, stack):
    if ref_col < 0 or ref_col >= len(raw_headers):
        raise ValueError("Header formula column out of range")

    value = raw_headers[ref_col]
    if value is None:
        return ""
    if isinstance(value, str) and value.startswith("="):
        if ref_col in stack:
            return ""
        return resolve_header_formula(value, raw_headers, stack | {ref_col})
    return str(value).strip()


def get_dataframe_cell_value(df, ref_col_label, ref_row_number):
    ref_col = column_index_from_string(ref_col_label) - 1
    ref_row_idx = int(ref_row_number) - 2
    if ref_col < 0 or ref_col >= len(df.columns):
        raise ValueError("Formula column out of range")
    if ref_row_idx < 0 or ref_row_idx >= len(df.index):
        raise ValueError("Formula row out of range")
    return df.iat[ref_row_idx, ref_col]


def resolve_formula_from_dataframe(formula, df, current_row_idx):
    expression = formula[1:]

    try:
        direct_ref_match = FULL_CELL_REF_PATTERN.fullmatch(expression.strip())
        if direct_ref_match:
            return get_dataframe_cell_value(df, direct_ref_match.group(1), direct_ref_match.group(2))

        if "&" in expression:
            parts = [part.strip() for part in expression.split("&")]
            resolved_parts = []

            for part in parts:
                ref_match = FULL_CELL_REF_PATTERN.fullmatch(part)
                if ref_match:
                    ref_value = get_dataframe_cell_value(df, ref_match.group(1), ref_match.group(2))
                    resolved_parts.append("" if pd.isna(ref_value) else str(ref_value))
                    continue
                if part.startswith('"') and part.endswith('"'):
                    resolved_parts.append(part[1:-1])
                    continue
                raise ValueError("Unsupported concatenation operand")

            return "".join(resolved_parts)

        def replace_cell(match):
            ref_value = get_dataframe_cell_value(df, match.group(1), match.group(2))
            if pd.isna(ref_value):
                return "0"

            return normalize_scalar(ref_value) or "0"

        arithmetic_expression = CELL_REF_PATTERN.sub(replace_cell, expression)
        if not ARITHMETIC_EXPRESSION_PATTERN.fullmatch(arithmetic_expression):
            raise ValueError("Unsupported formula expression")

        return safe_eval_arithmetic(arithmetic_expression)
    except Exception:
        logging.debug(
            "Unable to resolve formula %s on logical row %s",
            formula,
            current_row_idx + 2,
        )
        return None


def normalize_and_filter_rows(df, key_name, label):
    """
    Normalize sheet data for key-based comparison.

    Rows with a blank comparison key are not valid records for set comparison.
    This avoids false positives from Excel sheets whose trailing formula columns
    evaluate to values like 0 even when the business columns are empty.
    """
    df = df.copy()
    df.columns = [str(col).strip() for col in df.columns]

    for col in df.columns:
        df[col] = df[col].map(normalize_scalar)

    before = len(df)
    df = df[df[key_name] != ''].copy()
    removed = before - len(df)
    if removed:
        logging.info("Dropped %s rows with blank key from %s", removed, label)

    return df


def rows_match(row_s, row_t, formula_s, formula_t):
    columns = list(row_s.keys())
    columns.extend(col for col in row_t.keys() if col not in row_s)

    for col in columns:
        source_formula = formula_s.get(col, "")
        target_formula = formula_t.get(col, "")
        if source_formula and target_formula:
            if source_formula != target_formula:
                return False
            continue

        if row_s.get(col, "") != row_t.get(col, ""):
            return False

    return True


def cell_values_match(source_value, target_value, source_formula="", target_formula=""):
    if source_formula and target_formula:
        return source_formula == target_formula
    return normalize_scalar(source_value) == normalize_scalar(target_value)


def get_args():
    parser = argparse.ArgumentParser(
        description="Compare two Excel sheets and categorize differences into three files.")
    parser.add_argument("source_file", help="Path to source Excel file")
    parser.add_argument("source_sheet", help="Source sheet name")
    parser.add_argument("target_file", help="Path to target Excel file")
    parser.add_argument("target_sheet", help="Target sheet name")
    parser.add_argument("key_col", help="The Column letter to use as the Unique Key (e.g., A)")
    # Using 'store_true' means if you don't type --highlight, it is False.
    parser.add_argument("--highlight", action="store_true", help="Enable yellow highlighting for cell mismatches")
    return parser.parse_args()


def compare_excels():
    args = get_args()
    key_col_letter = args.key_col.upper()
    output_prefix = f"{args.source_sheet}_"

    # 1. Load data
    logging.info(f"Loading Source: {args.source_file}")
    df_s, formula_s = load_excel_sheet_data(args.source_file, args.source_sheet)

    logging.info(f"Loading Target: {args.target_file}")
    df_t, formula_t = load_excel_sheet_data(args.target_file, args.target_sheet)

    col_idx = column_index_from_string(key_col_letter) - 1
    key_name_s = df_s.columns[col_idx]
    key_name_t = df_t.columns[col_idx]

    source_row_numbers = pd.Series(range(2, len(df_s) + 2), index=df_s.index)
    target_row_numbers = pd.Series(range(2, len(df_t) + 2), index=df_t.index)
    df_s[key_name_s] = load_physical_column_values(
        args.source_file,
        args.source_sheet,
        key_col_letter,
        len(df_s),
    )
    df_t[key_name_t] = load_physical_column_values(
        args.target_file,
        args.target_sheet,
        key_col_letter,
        len(df_t),
    )

    df_s = normalize_and_filter_rows(df_s, key_name_s, "Source")
    df_t = normalize_and_filter_rows(df_t, key_name_t, "Target")
    source_row_numbers = source_row_numbers.loc[df_s.index].copy()
    target_row_numbers = target_row_numbers.loc[df_t.index].copy()
    formula_s = formula_s.loc[df_s.index].copy()
    formula_t = formula_t.loc[df_t.index].copy()
    formula_s.index = df_s[key_name_s]
    formula_t.index = df_t[key_name_t]
    source_row_numbers.index = df_s[key_name_s]
    target_row_numbers.index = df_t[key_name_t]

    # Set indices
    df_s.set_index(key_name_s, inplace=True, drop=False)
    df_t.set_index(key_name_t, inplace=True, drop=False)

    # 2. File 1: In Source Not In Target
    in_s_not_t = df_s[~df_s.index.isin(df_t.index)]
    in_source_not_in_target_file = f"{output_prefix}InSourceNotInTarget.xlsx"
    in_s_not_t.to_excel(in_source_not_in_target_file, index=False)
    logging.info(f"Saved {in_source_not_in_target_file} ({len(in_s_not_t)} records)")

    # 3. File 2: In Target Not In Source
    in_t_not_s = df_t[~df_t.index.isin(df_s.index)]
    in_target_not_in_source_file = f"{output_prefix}InTargetNotInSource.xlsx"
    in_t_not_s.to_excel(in_target_not_in_source_file, index=False)
    logging.info(f"Saved {in_target_not_in_source_file} ({len(in_t_not_s)} records)")

    # 4. File 3: In Both but Data Doesn't Match
    logging.info("Comparing overlapping keys...")
    common_keys = df_s.index[df_s.index.isin(df_t.index)].unique()

    # Use the first row per key as the comparison baseline on each side.
    # This keeps the comparison predictable even when the key is not unique.
    lookup_s = df_s[~df_s.index.duplicated(keep='first')].to_dict('index')
    lookup_t = df_t[~df_t.index.duplicated(keep='first')].to_dict('index')
    formula_lookup_s = formula_s[~formula_s.index.duplicated(keep='first')].to_dict('index')
    formula_lookup_t = formula_t[~formula_t.index.duplicated(keep='first')].to_dict('index')
    source_row_lookup = source_row_numbers[~source_row_numbers.index.duplicated(keep='first')].to_dict()
    target_row_lookup = target_row_numbers[~target_row_numbers.index.duplicated(keep='first')].to_dict()

    mismatched_rows = []

    for key in common_keys:
        row_s_dict = lookup_s.get(key)
        row_t_dict = lookup_t.get(key)
        formula_s_dict = formula_lookup_s.get(key, {})
        formula_t_dict = formula_lookup_t.get(key, {})

        if row_t_dict and not rows_match(row_s_dict, row_t_dict, formula_s_dict, formula_t_dict):
            mismatch_row = dict(row_s_dict)
            mismatch_row["_SourceExcelRow"] = source_row_lookup.get(key, "")
            mismatch_row["_TargetExcelRow"] = target_row_lookup.get(key, "")
            mismatch_row["_TargetKey"] = row_t_dict.get(key_name_t, "")
            mismatched_rows.append(mismatch_row)
            if len(mismatched_rows) >= MISMATCH_OUTPUT_THRESHOLD:
                logging.info(
                    "Mismatch threshold reached (%s records). Stopping early.",
                    MISMATCH_OUTPUT_THRESHOLD,
                )
                break

    mismatch_file = f"{output_prefix}InSourceAndInTarget.xlsx"
    mismatch_columns = list(df_s.columns) + ["_SourceExcelRow", "_TargetExcelRow", "_TargetKey"]
    if len(mismatched_rows) >= MISMATCH_OUTPUT_THRESHOLD:
        df_mismatch = pd.DataFrame(mismatched_rows, columns=mismatch_columns)
        logging.info(f"Saved {mismatch_file} ({len(mismatched_rows)} records)")
    else:
        df_mismatch = pd.DataFrame(columns=mismatch_columns)
        logging.info(
            "Mismatch count below threshold (%s). Writing header-only %s.",
            MISMATCH_OUTPUT_THRESHOLD,
            mismatch_file,
        )

    df_mismatch.to_excel(mismatch_file, index=False)

    if len(mismatched_rows) >= MISMATCH_OUTPUT_THRESHOLD and args.highlight:
        logging.info("Starting cell-level highlighting...")
        highlight_mismatches(mismatch_file, lookup_t, key_name_s, formula_lookup_s, formula_lookup_t)


def highlight_mismatches(output_file, lookup_dict, key_name, source_formula_lookup=None, target_formula_lookup=None):
    source_formula_lookup = source_formula_lookup or {}
    target_formula_lookup = target_formula_lookup or {}

    try:
        wb = load_workbook(output_file)
        ws = wb.active
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        headers = [cell.value for cell in ws[1]]
        key_idx = headers.index(key_name)

        for r_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            key_val = str(row[key_idx].value or "").strip()
            other_row = lookup_dict.get(key_val)
            source_formulas = source_formula_lookup.get(key_val, {})
            target_formulas = target_formula_lookup.get(key_val, {})

            if other_row:
                for c_idx, cell in enumerate(row):
                    col_name = headers[c_idx]
                    if not cell_values_match(
                        cell.value,
                        other_row.get(col_name, ""),
                        source_formulas.get(col_name, ""),
                        target_formulas.get(col_name, ""),
                    ):
                        cell.fill = yellow_fill

            if r_idx % 1000 == 0:
                logging.info(f"Processed {r_idx} rows...")

        wb.save(output_file)
        logging.info("Highlighting saved.")
    except Exception as e:
        logging.error(f"Highlighting failed: {e}")


if __name__ == "__main__":
    compare_excels()
