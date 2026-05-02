"""
FX Consolidation Post-Processing Tool
-------------------------------------
Purpose:
    Rebuilds 数据透视表 and 1数透结果 from the prepared FX workbook after
    the user has finished manual lookup inputs, then publishes the final report.
"""

import argparse
import logging
import os
from datetime import datetime, date

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill


LOG_FORMAT = '%(asctime)s - %(levelname)s - %(message)s'
DATE_FORMAT = '%H:%M:%S'
IN_PROGRESS_SUFFIX = '-处理中.xlsx'
SETTLEMENT_FLOW_INPUT_SHEET_NAME = '数据透视表'
SETTLEMENT_FLOW_OUTPUT_SHEET_NAME = '1数透结果'
CHANNEL_ORDER_SHEET_NAME = '渠道订单'
ACCOUNT_STATEMENT_SHEET_NAME = '账户流水'
DAILY_EXCHANGE_RATE_SHEET_NAME = '每日汇率(oc系统中获取）'
ESTIMATED_FX_SUMMARY_SHEET_NAME = '预估换汇汇总'
SUMMARY_SHEET_NAME = '处理摘要'
PIVOT_TOTAL_FILL = PatternFill(fill_type='solid', fgColor='FFF2F1F7')
PIVOT_HIGHLIGHT_FILL = PatternFill(fill_type='solid', fgColor='FFFFFF00')
RESULT_HEADER_FILL = PatternFill(fill_type='solid', fgColor='FF00B050')
FX_TRANSACTION_TITLE_FILL = PatternFill(fill_type='solid', fgColor='FFFFFF00')
FX_TRANSACTION_HEADER_FILL = PatternFill(fill_type='solid', fgColor='FFF2F1F7')
FX_TRANSACTION_NET_FILL = PatternFill(fill_type='solid', fgColor='FFD9D9D9')
FX_TRANSACTION_NET_RESULT_FILL = PatternFill(fill_type='solid', fgColor='FFFFC7CE')
GRAND_TOTAL_NUMBER_FORMAT = '0.00'
ESTIMATED_FX_SUMMARY_ACCOUNTING_FORMAT = '_(* #,##0.00_);_(* (#,##0.00);_(* "-"??_);_(@_)'
ESTIMATED_FX_SUMMARY_NUMBER_4_FORMAT = '0.0000'
ESTIMATED_FX_SUMMARY_NUMBER_2_FORMAT = '0.00'
SETTLEMENT_FLOW_INPUT_HEADERS = [
    '支付币种',
    '支付金额（扣除通道成本）',
    '打款币种',
    '清算币种',
    '清算金额',
    '打款币种与清算币种是否一致',
    '通道名称',
]
SETTLEMENT_FLOW_OUTPUT_HEADERS = [
    '支付币种',
    '打款币种',
    '清算币种',
    '求和项:支付金额（扣除通道成本）',
    '求和项:清算金额',
]


logging.basicConfig(
    level=logging.INFO,
    format=LOG_FORMAT,
    datefmt=DATE_FORMAT,
)


#######################################################
#  Common Utils
#######################################################
def configure_run_logging(log_path):
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)

    for handler in list(logger.handlers):
        if getattr(handler, '_fx_consolidation_postprocess_file_handler', False):
            logger.removeHandler(handler)
            handler.close()

    file_handler = logging.FileHandler(log_path, encoding='utf-8')
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(logging.Formatter(LOG_FORMAT, datefmt=DATE_FORMAT))
    file_handler._fx_consolidation_postprocess_file_handler = True
    logger.addHandler(file_handler)


def build_final_report_path(workbook_path):
    if workbook_path.endswith(IN_PROGRESS_SUFFIX):
        return workbook_path[:-len(IN_PROGRESS_SUFFIX)] + '.xlsx'
    return workbook_path


def publish_final_report(workbook_path):
    final_path = build_final_report_path(workbook_path)
    if final_path == workbook_path:
        return final_path
    if os.path.exists(final_path):
        os.remove(final_path)
    os.rename(workbook_path, final_path)
    return final_path


def normalize_cell_text(value):
    if value is None:
        return ""
    return str(value).strip()


def numeric_cell_value(value, default=0.0):
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if text == "":
        return default

    text = text.replace(",", "").replace("$", "")
    try:
        return float(text)
    except ValueError:
        return default


def to_excel_cell_value(value):
    if value is None:
        return None
    if isinstance(value, str) and value == "":
        return None
    return value


def normalize_comparable_value(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def parse_account_statement_datetime(value):
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())

    text = normalize_cell_text(value)
    if text == "":
        return None

    supported_formats = (
        '%d/%m/%Y %H:%M:%S',
        '%Y-%m-%d %H:%M:%S',
        '%Y-%m-%d',
        '%d/%m/%Y',
    )
    for fmt in supported_formats:
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue

    raise ValueError(f"Failed to parse 账户流水 datetime value: {text}")


def write_matrix(ws, start_row, start_col, rows):
    for row_offset, row_values in enumerate(rows):
        for col_offset, value in enumerate(row_values):
            ws.cell(row=start_row + row_offset, column=start_col + col_offset).value = to_excel_cell_value(value)


def write_header_row(ws, start_col, headers, fill):
    for col_offset, header in enumerate(headers):
        cell = ws.cell(row=1, column=start_col + col_offset)
        cell.value = header
        cell.fill = fill


def get_last_data_row(ws, min_col, max_col, min_row=1):
    for row_number in range(ws.max_row, min_row - 1, -1):
        for col_number in range(min_col, max_col + 1):
            if normalize_cell_text(ws.cell(row=row_number, column=col_number).value) != "":
                return row_number
    return min_row - 1


def get_range_values(ws, start_row, end_row, start_col, end_col):
    values = []
    if end_row < start_row:
        return values

    for row_number in range(start_row, end_row + 1):
        row_values = []
        for col_number in range(start_col, end_col + 1):
            row_values.append(ws.cell(row=row_number, column=col_number).value)
        values.append(row_values)
    return values


def build_excel_formula(formula_body):
    return f"={formula_body}"


def round_to_nearest_hundred(value):
    numeric_value = numeric_cell_value(value)
    if numeric_value >= 0:
        return int((numeric_value + 50) // 100 * 100)
    return -int((abs(numeric_value) + 50) // 100 * 100)


def build_lookup_map_from_sheet(ws, key_col_idx, value_col_idx):
    lookup = {}
    for row_number in range(2, ws.max_row + 1):
        key = normalize_cell_text(ws.cell(row=row_number, column=key_col_idx).value)
        if key == "":
            continue
        lookup[key] = ws.cell(row=row_number, column=value_col_idx).value
    return lookup


def update_first_match_lookup(lookup, key, value):
    if key == "":
        return

    # Mirror Excel XLOOKUP's default top-to-bottom behavior: keep the first
    # match, but allow a later nonblank value to replace an initially blank one.
    if key not in lookup:
        lookup[key] = value
    elif normalize_cell_text(lookup[key]) == "" and normalize_cell_text(value) != "":
        lookup[key] = value


def build_payout_currency_lookup(ws_payout_currency):
    exact_lookup = {}
    a07_relaxed_lookup = {}
    channel_entity_currency_lookup = {}

    for row_number in range(2, ws_payout_currency.max_row + 1):
        channel_name = normalize_cell_text(ws_payout_currency.cell(row=row_number, column=2).value)
        entity_value = normalize_cell_text(ws_payout_currency.cell(row=row_number, column=3).value)
        merchant_value = normalize_cell_text(ws_payout_currency.cell(row=row_number, column=4).value)
        currency_value = normalize_cell_text(ws_payout_currency.cell(row=row_number, column=5).value)
        key = "".join([
            channel_name,
            entity_value,
            merchant_value,
            currency_value,
        ])
        value = ws_payout_currency.cell(row=row_number, column=6).value

        # Exact lookup mirrors 打款币种!G -> F, where G is effectively B+C+D+E.
        update_first_match_lookup(exact_lookup, key, value)

        if channel_name == "A07":
            # Some saved workbooks contain A07 rows where the effective lookup key
            # behaves like B+D+E rather than the full B+C+D+E formula chain.
            update_first_match_lookup(
                a07_relaxed_lookup,
                f"{channel_name}{merchant_value}{currency_value}",
                value,
            )

        if channel_name:
            # Some channel rows contain verbose text in 打款币种!D. This fallback
            # uses B+C+E so the postprocess can still recover the payout currency.
            update_first_match_lookup(
                channel_entity_currency_lookup,
                f"{channel_name}{entity_value}{currency_value}",
                value,
            )

    return {
        'exact_lookup': exact_lookup,
        'a07_relaxed_lookup': a07_relaxed_lookup,
        'channel_entity_currency_lookup': channel_entity_currency_lookup,
    }


def resolve_settlement_flow_aq_value(ap_value, ah_value, ab_value, a01_lookup, a07_lookup):
    ap_value_text = normalize_cell_text(ap_value)
    ah_value_text = normalize_cell_text(ah_value)
    if ap_value_text == "2号通道":
        return a01_lookup.get(ah_value_text, "")
    if ap_value_text == "A07":
        return f"{normalize_cell_text(a07_lookup.get(ah_value_text, ''))}{ah_value_text}" if ah_value_text else ""
    if ap_value_text == "7号通道":
        return ab_value
    return ""


def build_settlement_flow_ar_candidates(ap_value, aq_value, ah_value, aj_value):
    ap_value_text = normalize_cell_text(ap_value)
    aq_value_text = normalize_cell_text(aq_value)
    ah_value_text = normalize_cell_text(ah_value)
    aj_value_text = normalize_cell_text(aj_value)

    candidates = []

    def add_candidate(value):
        if value and value not in candidates:
            candidates.append(value)

    # First candidate mirrors the 渠道订单 formula chain exactly:
    # AR = AP & AQ & AJ.
    add_candidate(f"{ap_value_text}{aq_value_text}{aj_value_text}")
    if ap_value_text == "A07" and ah_value_text:
        # Some saved workbooks behave as if A07 uses AP & AH & AJ directly when
        # resolving 打款币种, so keep that as a compatibility fallback.
        add_candidate(f"{ap_value_text}{ah_value_text}{aj_value_text}")

    return candidates


def build_account_statement_lookup(ws_account_statement):
    lookup = {}

    for row_number in range(2, ws_account_statement.max_row + 1):
        col_b = normalize_cell_text(ws_account_statement.cell(row=row_number, column=2).value)
        col_e = normalize_cell_text(ws_account_statement.cell(row=row_number, column=5).value)
        internal_key = f"{col_b}{col_e}"
        if internal_key == "" or internal_key in lookup:
            continue

        lookup[internal_key] = {
            'am_value': ws_account_statement.cell(row=row_number, column=7).value,
            'an_value': numeric_cell_value(ws_account_statement.cell(row=row_number, column=11).value)
            - numeric_cell_value(ws_account_statement.cell(row=row_number, column=12).value),
        }

    return lookup


def build_settlement_flow_context(write_wb):
    return {
        'channel_name_lookup': build_lookup_map_from_sheet(write_wb['渠道名称'], 1, 2),
        'payout_currency_lookup': build_payout_currency_lookup(write_wb['打款币种']),
        'a01_lookup': build_lookup_map_from_sheet(write_wb['二级商户号映射表-A01'], 1, 2),
        'a07_lookup': build_lookup_map_from_sheet(write_wb['二级商户号映射表-A07'], 1, 3),
    }


#######################################################
#  数据透视表 Creation and Input Rebuild
#######################################################
def recreate_pivot_source_sheet(wb):
    if SETTLEMENT_FLOW_INPUT_SHEET_NAME in wb.sheetnames:
        sheet_index = wb.sheetnames.index(SETTLEMENT_FLOW_INPUT_SHEET_NAME)
        old_sheet = wb[SETTLEMENT_FLOW_INPUT_SHEET_NAME]
        wb.remove(old_sheet)
    else:
        sheet_index = len(wb.sheetnames)

    ws_pivot = wb.create_sheet(SETTLEMENT_FLOW_INPUT_SHEET_NAME, sheet_index)
    write_header_row(ws_pivot, 1, SETTLEMENT_FLOW_INPUT_HEADERS, PIVOT_HIGHLIGHT_FILL)
    write_header_row(ws_pivot, 11, SETTLEMENT_FLOW_OUTPUT_HEADERS, PIVOT_HIGHLIGHT_FILL)
    return ws_pivot


def build_pivot_source_rows(ws_channel_orders, account_statement_lookup, settlement_flow_context, ws_channel_order_values=None):
    pivot_rows = []
    cached_blank_al_count = 0
    cached_blank_al_samples = []
    fallback_blank_al_count = 0
    fallback_blank_al_samples = []

    for row_number in range(2, ws_channel_orders.max_row + 1):
        if normalize_cell_text(ws_channel_orders.cell(row=row_number, column=1).value) == "":
            continue

        if ws_channel_order_values is not None:
            cached_ao_value = normalize_cell_text(ws_channel_order_values.cell(row=row_number, column=41).value)
            if cached_ao_value in {"是", "否"}:
                if cached_ao_value == "否":
                    cached_al_value = ws_channel_order_values.cell(row=row_number, column=38).value
                    if normalize_cell_text(cached_al_value) == "":
                        cached_blank_al_count += 1
                        if len(cached_blank_al_samples) < 10:
                            cached_blank_al_samples.append({
                                'row': row_number,
                                'aj': ws_channel_order_values.cell(row=row_number, column=36).value,
                                'al': cached_al_value,
                                'am': ws_channel_order_values.cell(row=row_number, column=39).value,
                                'ao': cached_ao_value,
                                'ap': ws_channel_order_values.cell(row=row_number, column=42).value,
                                'aq': ws_channel_order_values.cell(row=row_number, column=43).value,
                                'ar': ws_channel_order_values.cell(row=row_number, column=44).value,
                                'channel_code': ws_channel_orders.cell(row=row_number, column=6).value,
                                'merchant_no': ws_channel_orders.cell(row=row_number, column=4).value,
                                'merchant_order_no': ws_channel_orders.cell(row=row_number, column=5).value,
                            })
                    pivot_rows.append([
                        ws_channel_order_values.cell(row=row_number, column=36).value,
                        ws_channel_order_values.cell(row=row_number, column=37).value,
                        cached_al_value,
                        ws_channel_order_values.cell(row=row_number, column=39).value,
                        ws_channel_order_values.cell(row=row_number, column=40).value,
                        ws_channel_order_values.cell(row=row_number, column=41).value,
                        ws_channel_order_values.cell(row=row_number, column=42).value,
                    ])
                continue

        col_d = normalize_cell_text(ws_channel_orders.cell(row=row_number, column=4).value)
        col_e = normalize_cell_text(ws_channel_orders.cell(row=row_number, column=5).value)
        internal_key = f"{col_d}{col_e}"
        if internal_key not in account_statement_lookup:
            continue

        col_f = normalize_cell_text(ws_channel_orders.cell(row=row_number, column=6).value)
        col_i = normalize_cell_text(ws_channel_orders.cell(row=row_number, column=9).value)
        col_m = numeric_cell_value(ws_channel_orders.cell(row=row_number, column=13).value)
        col_n = ws_channel_orders.cell(row=row_number, column=14).value
        col_ab = ws_channel_orders.cell(row=row_number, column=28).value
        col_ah = ws_channel_orders.cell(row=row_number, column=34).value

        ap_value = settlement_flow_context['channel_name_lookup'].get(col_f, "")
        aq_value = resolve_settlement_flow_aq_value(
            ap_value,
            col_ah,
            col_ab,
            settlement_flow_context['a01_lookup'],
            settlement_flow_context['a07_lookup'],
        )
        ak_value = -col_m * (1 - 0.032) if col_i == "退款" else col_m * (1 - 0.032)
        ar_candidates = build_settlement_flow_ar_candidates(ap_value, aq_value, col_ah, col_n)
        ar_value = ar_candidates[0] if ar_candidates else ""
        payout_lookup = settlement_flow_context['payout_currency_lookup']
        al_value = ""
        matched_ar_value = ""

        # Resolution order:
        # 1. Exact AR match against the normalized 打款币种 key space.
        # 2. A07-specific compatibility fallback keyed by A07 + AH + AJ.
        # 3. Channel/entity/currency fallback for rows whose 打款币种!D contains
        #    descriptive text that should not participate in the lookup key.
        for candidate in ar_candidates:
            candidate_value = payout_lookup['exact_lookup'].get(candidate, "")
            if normalize_cell_text(candidate_value) != "":
                al_value = candidate_value
                matched_ar_value = candidate
                break
        if normalize_cell_text(al_value) == "" and normalize_cell_text(ap_value) == "A07":
            relaxed_candidate = f"A07{normalize_cell_text(col_ah)}{normalize_cell_text(col_n)}"
            candidate_value = payout_lookup['a07_relaxed_lookup'].get(relaxed_candidate, "")
            if normalize_cell_text(candidate_value) != "":
                al_value = candidate_value
                matched_ar_value = relaxed_candidate
        if normalize_cell_text(al_value) == "" and normalize_cell_text(ap_value) == "2号通道":
            relaxed_candidate = f"{normalize_cell_text(ap_value)}{normalize_cell_text(aq_value)}{normalize_cell_text(col_n)}"
            candidate_value = payout_lookup['channel_entity_currency_lookup'].get(relaxed_candidate, "")
            if normalize_cell_text(candidate_value) != "":
                al_value = candidate_value
                matched_ar_value = relaxed_candidate
        if matched_ar_value:
            ar_value = matched_ar_value
        am_value = account_statement_lookup[internal_key]['am_value']
        an_value = account_statement_lookup[internal_key]['an_value']
        ao_value = "是" if normalize_cell_text(al_value) == normalize_cell_text(am_value) else "否"
        if ao_value != "否":
            continue

        if normalize_cell_text(al_value) == "":
            fallback_blank_al_count += 1
            if len(fallback_blank_al_samples) < 10:
                fallback_blank_al_samples.append({
                    'row': row_number,
                    'aj': col_n,
                    'al': al_value,
                    'am': am_value,
                    'ao': ao_value,
                    'ap': ap_value,
                    'aq': aq_value,
                    'ar': ar_value,
                    'ar_candidates': ar_candidates,
                    'channel_code': col_f,
                    'merchant_no': col_d,
                    'merchant_order_no': col_e,
                    'ah': col_ah,
                    'ab': col_ab,
                })

        pivot_rows.append([
            col_n,
            ak_value,
            al_value,
            am_value,
            an_value,
            ao_value,
            ap_value,
        ])

    if cached_blank_al_count:
        logging.warning(
            "Cached 渠道订单 rows with AO=否 but blank AL: count=%s sample=%s",
            cached_blank_al_count,
            cached_blank_al_samples,
        )
    if fallback_blank_al_count:
        logging.warning(
            "Fallback-rebuilt 渠道订单 rows with AO=否 but blank AL: count=%s sample=%s",
            fallback_blank_al_count,
            fallback_blank_al_samples,
        )

    return pivot_rows


def validate_settlement_flow_input_sheet(ws_pivot, expected_rows):
    actual_last_row = get_last_data_row(ws_pivot, 1, 7, min_row=2)
    actual_rows = get_range_values(ws_pivot, 2, actual_last_row, 1, 7) if actual_last_row >= 2 else []
    normalized_actual = [[normalize_comparable_value(value) for value in row] for row in actual_rows]
    normalized_expected = [[normalize_comparable_value(value) for value in row] for row in expected_rows]

    if len(actual_rows) != len(expected_rows):
        raise ValueError(
            f"Settlement flow input validation failed: expected {len(expected_rows)} rows, "
            f"found {len(actual_rows)} rows in 数据透视表!A:G."
        )
    if normalized_actual != normalized_expected:
        raise ValueError(
            "Settlement flow input validation failed: 数据透视表!A:G values do not match "
            "the expected AJ:AP value mapping from 渠道订单."
        )


def rebuild_settlement_flow_input_sheet(write_wb, read_wb=None):
    logging.info("Rebuilding 数据透视表 input area A:G from 渠道订单 and 账户流水...")
    ws_pivot = recreate_pivot_source_sheet(write_wb)
    ws_channel_order_values = None
    if read_wb is not None and CHANNEL_ORDER_SHEET_NAME in read_wb.sheetnames:
        ws_channel_order_values = read_wb[CHANNEL_ORDER_SHEET_NAME]
    pivot_rows = build_pivot_source_rows(
        write_wb[CHANNEL_ORDER_SHEET_NAME],
        build_account_statement_lookup(write_wb[ACCOUNT_STATEMENT_SHEET_NAME]),
        build_settlement_flow_context(write_wb),
        ws_channel_order_values,
    )
    write_matrix(ws_pivot, 2, 1, pivot_rows)
    validate_settlement_flow_input_sheet(ws_pivot, pivot_rows)

    logging.info("Settlement flow input validated: rebuilt %s rows in 数据透视表!A:G", len(pivot_rows))
    return ws_pivot, pivot_rows


#######################################################
#  数据透视表 Summary Build
#######################################################
def build_grouped_pivot_rows(source_rows):
    grouped = {}
    for row in source_rows:
        group_key = (
            normalize_cell_text(row[0]),
            normalize_cell_text(row[2]),
            normalize_cell_text(row[3]),
        )
        bucket = grouped.setdefault(group_key, {'sum_b': 0.0, 'sum_e': 0.0})
        bucket['sum_b'] += numeric_cell_value(row[1])
        bucket['sum_e'] += numeric_cell_value(row[4])

    grouped_rows = []
    for group_key in sorted(grouped.keys()):
        totals = grouped[group_key]
        grouped_rows.append([
            group_key[0],
            group_key[1],
            group_key[2],
            totals['sum_b'],
            totals['sum_e'],
        ])

    grand_total_n = sum(row[3] for row in grouped_rows)
    grand_total_o = sum(row[4] for row in grouped_rows)
    grand_total_row = ["Grand Total", None, None, grand_total_n, grand_total_o]
    return grouped_rows, grand_total_row


def validate_settlement_flow_summary(ws_pivot, grouped_rows, grand_total_row):
    actual_last_row = get_last_data_row(ws_pivot, 11, 15, min_row=2)
    actual_rows = get_range_values(ws_pivot, 2, actual_last_row, 11, 15) if actual_last_row >= 2 else []
    expected_rows = grouped_rows + [grand_total_row]
    normalized_actual = [[normalize_comparable_value(value) for value in row] for row in actual_rows]
    normalized_expected = [[normalize_comparable_value(value) for value in row] for row in expected_rows]

    if normalized_actual != normalized_expected:
        raise ValueError("Settlement flow summary validation failed: 数据透视表!K:O does not match the grouped summary output.")

    grand_total_matches = [idx for idx, row in enumerate(actual_rows, start=2) if normalize_cell_text(row[0]) == "Grand Total"]
    if len(grand_total_matches) != 1:
        raise ValueError("Settlement flow summary validation failed: expected exactly one Grand Total row in 数据透视表!K:O.")

    grand_total_row_number = grand_total_matches[0]
    for col_number in range(11, 16):
        cell = ws_pivot.cell(row=grand_total_row_number, column=col_number)
        if cell.fill.fill_type != PIVOT_TOTAL_FILL.fill_type or cell.fill.fgColor.rgb != PIVOT_TOTAL_FILL.fgColor.rgb:
            raise ValueError("Settlement flow summary validation failed: Grand Total fill was not applied correctly to K:O.")

    for col_number in (14, 15):
        cell = ws_pivot.cell(row=grand_total_row_number, column=col_number)
        if cell.number_format != GRAND_TOTAL_NUMBER_FORMAT:
            raise ValueError("Settlement flow summary validation failed: Grand Total values in N:O were not formatted to two decimals.")


def build_settlement_flow_summary(ws_pivot, source_rows):
    logging.info("Building 数据透视表 summary table in K:O from %s source rows...", len(source_rows))
    grouped_rows, grand_total_row = build_grouped_pivot_rows(source_rows)

    for row_number in range(2, ws_pivot.max_row + 1):
        for col_number in range(11, 16):
            ws_pivot.cell(row=row_number, column=col_number).value = None
            ws_pivot.cell(row=row_number, column=col_number).fill = PatternFill(fill_type=None)

    output_rows = grouped_rows + [grand_total_row]
    write_matrix(ws_pivot, 2, 11, output_rows)

    grand_total_row_number = len(grouped_rows) + 2
    for col_number in range(11, 16):
        ws_pivot.cell(row=grand_total_row_number, column=col_number).fill = PIVOT_TOTAL_FILL
    for col_number in (14, 15):
        ws_pivot.cell(row=grand_total_row_number, column=col_number).number_format = GRAND_TOTAL_NUMBER_FORMAT

    validate_settlement_flow_summary(ws_pivot, grouped_rows, grand_total_row)
    logging.info(
        "Settlement flow summary validated: wrote %s grouped rows plus Grand Total to 数据透视表!K:O",
        len(grouped_rows),
    )


#######################################################
#  1数透结果 Publish
#######################################################
def validate_settlement_flow_results(ws_result, source_rows, remapped_rows, highlighted_row_count):
    actual_ae_last_row = get_last_data_row(ws_result, 1, 5, min_row=1)
    actual_ae_rows = get_range_values(ws_result, 1, actual_ae_last_row, 1, 5) if actual_ae_last_row >= 1 else []
    normalized_actual_ae = [[normalize_comparable_value(value) for value in row] for row in actual_ae_rows]
    normalized_expected_ae = [[normalize_comparable_value(value) for value in row] for row in source_rows]
    if normalized_actual_ae != normalized_expected_ae:
        raise ValueError("Settlement flow result validation failed: 1数透结果!A:E does not match 数据透视表!K:O.")

    if highlighted_row_count > 1:
        raise ValueError("Settlement flow result validation failed: more than one CNY/USD/CNY row was highlighted.")

    actual_hl_last_row = get_last_data_row(ws_result, 8, 12, min_row=1)
    actual_hl_rows = get_range_values(ws_result, 1, actual_hl_last_row, 8, 12) if actual_hl_last_row >= 1 else []
    normalized_actual_hl = [[normalize_comparable_value(value) for value in row] for row in actual_hl_rows]
    normalized_expected_hl = [[normalize_comparable_value(value) for value in row] for row in remapped_rows]
    if normalized_actual_hl != normalized_expected_hl:
        raise ValueError("Settlement flow result validation failed: 1数透结果!H:L does not match the filtered/remapped copy.")


def publish_settlement_flow_results(write_wb, ws_pivot):
    if SETTLEMENT_FLOW_OUTPUT_SHEET_NAME not in write_wb.sheetnames:
        raise KeyError(f"Target sheet not found in workbook: {SETTLEMENT_FLOW_OUTPUT_SHEET_NAME}")

    ws_result = write_wb[SETTLEMENT_FLOW_OUTPUT_SHEET_NAME]
    logging.info("Publishing 1数透结果 tables from 数据透视表: A:E primary copy and H:L remapped copy...")
    source_last_row = get_last_data_row(ws_pivot, 11, 15, min_row=1)
    source_rows = get_range_values(ws_pivot, 1, source_last_row, 11, 15) if source_last_row >= 1 else []

    if ws_result.max_row > 0:
        ws_result.delete_rows(1, ws_result.max_row)

    write_matrix(ws_result, 1, 1, source_rows)
    for col_number in range(1, 6):
        ws_result.cell(row=1, column=col_number).fill = RESULT_HEADER_FILL

    highlighted_row_count = 0
    remapped_rows = []
    for row_index, row_values in enumerate(source_rows, start=1):
        col_a = normalize_cell_text(row_values[0]) if len(row_values) > 0 else ""
        col_b = normalize_cell_text(row_values[1]) if len(row_values) > 1 else ""
        col_c = normalize_cell_text(row_values[2]) if len(row_values) > 2 else ""

        is_grand_total = col_a == "Grand Total"
        is_target_highlight = col_a == "CNY" and col_b == "USD" and col_c == "CNY"

        if is_target_highlight:
            highlighted_row_count += 1
            for col_number in range(1, 6):
                ws_result.cell(row=row_index, column=col_number).fill = PIVOT_HIGHLIGHT_FILL

        if row_index == 1 or (not is_target_highlight and not is_grand_total):
            remapped_rows.append([
                row_values[0] if len(row_values) > 0 else None,
                row_values[3] if len(row_values) > 3 else None,
                row_values[1] if len(row_values) > 1 else None,
                row_values[2] if len(row_values) > 2 else None,
                row_values[4] if len(row_values) > 4 else None,
            ])

    write_matrix(ws_result, 1, 8, remapped_rows)
    for col_number in range(8, 13):
        ws_result.cell(row=1, column=col_number).fill = RESULT_HEADER_FILL

    validate_settlement_flow_results(ws_result, source_rows, remapped_rows, highlighted_row_count)
    logging.info(
        "Settlement flow results validated: copied %s rows to 1数透结果!A:E and %s rows to 1数透结果!H:L",
        len(source_rows),
        len(remapped_rows),
    )


#######################################################
#  账户流水 / 预估换汇汇总 Transaction Date Derivation
#######################################################
def build_transaction_dates(ws_account_statement):
    logging.info("Deriving transaction date label from 账户流水 column A...")
    unique_dates = set()

    for row_number in range(2, ws_account_statement.max_row + 1):
        parsed_datetime = parse_account_statement_datetime(ws_account_statement.cell(row=row_number, column=1).value)
        if parsed_datetime is not None:
            unique_dates.add(parsed_datetime.date())

    if not unique_dates:
        raise ValueError("No valid transaction dates were found in 账户流水!A.")

    sorted_dates = sorted(unique_dates)
    formatted_dates = [date_value.strftime('%Y-%m-%d') for date_value in sorted_dates]
    if len(formatted_dates) == 1:
        transaction_dates = formatted_dates[0]
    else:
        transaction_dates = formatted_dates[0] + ''.join(f"&{date_value[-2:]}" for date_value in formatted_dates[1:])

    logging.info("Derived transaction date label: %s (%s unique dates)", transaction_dates, len(formatted_dates))
    return transaction_dates


#######################################################
#  每日汇率(oc系统中获取） Lookup Preparation
#######################################################
def build_daily_exchange_rate_lookup(ws_daily_exchange_rate):
    logging.info("Building 每日汇率(oc系统中获取） lookup from column I/H with D/E fallback keys...")
    exchange_rate_lookup = {}

    for row_number in range(2, ws_daily_exchange_rate.max_row + 1):
        key = normalize_cell_text(ws_daily_exchange_rate.cell(row=row_number, column=9).value)
        if key == "":
            source_currency = normalize_cell_text(ws_daily_exchange_rate.cell(row=row_number, column=4).value)
            target_currency = normalize_cell_text(ws_daily_exchange_rate.cell(row=row_number, column=5).value)
            key = f"{source_currency}{target_currency}"
        if key == "":
            continue
        update_first_match_lookup(
            exchange_rate_lookup,
            key,
            numeric_cell_value(ws_daily_exchange_rate.cell(row=row_number, column=8).value),
        )

    logging.info("Built 每日汇率 lookup with %s keys", len(exchange_rate_lookup))
    return exchange_rate_lookup


#######################################################
#  1数透结果 / 预估换汇汇总 Row Calculation
#######################################################
def build_estimated_fx_summary_rows(ws_settlement_flow_output, transaction_dates, exchange_rate_lookup):
    logging.info(
        "Calculating 预估换汇汇总 top table rows from 1数透结果!H:L using transaction label %s and %s exchange-rate keys...",
        transaction_dates,
        len(exchange_rate_lookup),
    )
    source_last_row = get_last_data_row(ws_settlement_flow_output, 8, 12, min_row=2)
    source_rows = get_range_values(ws_settlement_flow_output, 2, source_last_row, 8, 12) if source_last_row >= 2 else []
    estimated_rows = []

    for row_offset, source_row in enumerate(source_rows, start=2):
        b_value = source_row[0]
        c_value = numeric_cell_value(source_row[1])
        d_value = source_row[2]
        g_value = source_row[3]
        h_value = numeric_cell_value(source_row[4])
        b_value_text = normalize_cell_text(b_value)
        d_value_text = normalize_cell_text(d_value)
        lookup_key = f"{b_value_text}{d_value_text}"

        if b_value_text == "":
            raise ValueError(
                f"Blank payment currency in 1数透结果!H{row_offset}; "
                "cannot build 预估换汇汇总 Daily Exchange Rate lookup key."
            )
        if d_value_text == "":
            raise ValueError(
                f"Blank payout currency in 1数透结果!J{row_offset}; "
                f"cannot build Daily Exchange Rate lookup key from {b_value_text!r} and a blank payout currency."
            )

        if b_value_text == d_value_text:
            e_value = 1.0
        else:
            if lookup_key not in exchange_rate_lookup:
                raise ValueError(f"Missing Daily Exchange Rate lookup for key: {lookup_key}")
            e_value = exchange_rate_lookup[lookup_key]

        f_value = c_value * e_value
        i_value = h_value * (1 - 0.03)
        j_value = f"{normalize_cell_text(d_value)}{normalize_cell_text(g_value)}"

        estimated_rows.append({
            'excel_row': row_offset,
            'a_value': transaction_dates,
            'b_value': b_value,
            'c_value': c_value,
            'd_value': d_value,
            'e_value': e_value,
            'f_value': f_value,
            'g_value': g_value,
            'h_value': h_value,
            'i_value': i_value,
            'j_value': j_value,
            'e_formula': build_excel_formula(
                f"IF(B{row_offset}=D{row_offset},1,XLOOKUP(B{row_offset}&D{row_offset},'{DAILY_EXCHANGE_RATE_SHEET_NAME}'!I:I,'{DAILY_EXCHANGE_RATE_SHEET_NAME}'!H:H))"
            ),
            'f_formula': build_excel_formula(f"C{row_offset}*E{row_offset}"),
            'i_formula': build_excel_formula(f"H{row_offset}*(1-3%)"),
            'j_formula': build_excel_formula(f"D{row_offset}&G{row_offset}"),
        })

    logging.info("Calculated %s rows for 预估换汇汇总 top table", len(estimated_rows))
    return estimated_rows


#######################################################
#  预估换汇汇总 Publish
#######################################################
def apply_estimated_fx_summary_top_table_formats(ws_estimated_fx_summary, start_row, end_row):
    for row_number in range(start_row, end_row + 1):
        for col_number in (3, 8, 9):
            ws_estimated_fx_summary.cell(row=row_number, column=col_number).number_format = ESTIMATED_FX_SUMMARY_ACCOUNTING_FORMAT
        ws_estimated_fx_summary.cell(row=row_number, column=5).number_format = ESTIMATED_FX_SUMMARY_NUMBER_4_FORMAT
        ws_estimated_fx_summary.cell(row=row_number, column=6).number_format = ESTIMATED_FX_SUMMARY_NUMBER_2_FORMAT


def validate_estimated_fx_summary_output(ws_estimated_fx_summary, estimated_rows):
    actual_last_row = get_last_data_row(ws_estimated_fx_summary, 1, 10, min_row=2)
    actual_rows = get_range_values(ws_estimated_fx_summary, 2, actual_last_row, 1, 10) if actual_last_row >= 2 else []
    expected_total_row_number = len(estimated_rows) + 2

    if len(actual_rows) != len(estimated_rows) + 1:
        raise ValueError(
            f"Estimated FX Summary validation failed: expected {len(estimated_rows) + 1} rows, "
            f"found {len(actual_rows)} rows in 预估换汇汇总."
        )

    for actual_row, expected_row in zip(actual_rows[:-1], estimated_rows):
        expected_prefix = [
            expected_row['a_value'],
            expected_row['b_value'],
            expected_row['c_value'],
            expected_row['d_value'],
        ]
        if [normalize_comparable_value(value) for value in actual_row[:4]] != [normalize_comparable_value(value) for value in expected_prefix]:
            raise ValueError("Estimated FX Summary validation failed: values in A:D do not match the computed rows.")
        if normalize_comparable_value(actual_row[6]) != normalize_comparable_value(expected_row['g_value']) or normalize_comparable_value(actual_row[7]) != normalize_comparable_value(expected_row['h_value']):
            raise ValueError("Estimated FX Summary validation failed: values in G:H do not match the computed rows.")
        if actual_row[4] != expected_row['e_formula'] or actual_row[5] != expected_row['f_formula'] or actual_row[8] != expected_row['i_formula'] or actual_row[9] != expected_row['j_formula']:
            raise ValueError("Estimated FX Summary validation failed: formulas in E/F/I/J were not preserved correctly.")

    total_row = actual_rows[-1]
    if normalize_cell_text(total_row[0]) != 'Grand Total':
        raise ValueError("Estimated FX Summary validation failed: total row label in column A is incorrect.")
    if total_row[2] != build_excel_formula(f"SUM(C2:C{expected_total_row_number - 1})"):
        raise ValueError("Estimated FX Summary validation failed: total row formula in column C is incorrect.")
    if total_row[5] != build_excel_formula(f"SUM(F2:F{expected_total_row_number - 1})"):
        raise ValueError("Estimated FX Summary validation failed: total row formula in column F is incorrect.")
    if total_row[7] != build_excel_formula(f"SUM(H2:H{expected_total_row_number - 1})"):
        raise ValueError("Estimated FX Summary validation failed: total row formula in column H is incorrect.")
    if total_row[8] != build_excel_formula(f"SUM(I2:I{expected_total_row_number - 1})"):
        raise ValueError("Estimated FX Summary validation failed: total row formula in column I is incorrect.")

    for col_number in range(1, 11):
        cell = ws_estimated_fx_summary.cell(row=expected_total_row_number, column=col_number)
        if cell.fill.fill_type != FX_TRANSACTION_HEADER_FILL.fill_type or cell.fill.fgColor.rgb != FX_TRANSACTION_HEADER_FILL.fgColor.rgb:
            raise ValueError("Estimated FX Summary validation failed: total row fill was not applied correctly.")

    for row_number in range(2, expected_total_row_number + 1):
        for col_number in (3, 8, 9):
            if ws_estimated_fx_summary.cell(row=row_number, column=col_number).number_format != ESTIMATED_FX_SUMMARY_ACCOUNTING_FORMAT:
                raise ValueError("Estimated FX Summary validation failed: accounting format was not applied correctly.")
        if ws_estimated_fx_summary.cell(row=row_number, column=5).number_format != ESTIMATED_FX_SUMMARY_NUMBER_4_FORMAT:
            raise ValueError("Estimated FX Summary validation failed: number format in column E is incorrect.")
        if ws_estimated_fx_summary.cell(row=row_number, column=6).number_format != ESTIMATED_FX_SUMMARY_NUMBER_2_FORMAT:
            raise ValueError("Estimated FX Summary validation failed: number format in column F is incorrect.")


def publish_estimated_fx_summary(write_wb, estimated_rows):
    if ESTIMATED_FX_SUMMARY_SHEET_NAME not in write_wb.sheetnames:
        raise KeyError(f"Target sheet not found in workbook: {ESTIMATED_FX_SUMMARY_SHEET_NAME}")

    ws_estimated_fx_summary = write_wb[ESTIMATED_FX_SUMMARY_SHEET_NAME]
    logging.info("Publishing 预估换汇汇总 top table to A:J with %s detail rows plus Grand Total...", len(estimated_rows))
    if ws_estimated_fx_summary.max_row > 1:
        ws_estimated_fx_summary.delete_rows(2, ws_estimated_fx_summary.max_row - 1)

    for row_data in estimated_rows:
        row_number = row_data['excel_row']
        ws_estimated_fx_summary.cell(row=row_number, column=1).value = to_excel_cell_value(row_data['a_value'])
        ws_estimated_fx_summary.cell(row=row_number, column=2).value = to_excel_cell_value(row_data['b_value'])
        ws_estimated_fx_summary.cell(row=row_number, column=3).value = to_excel_cell_value(row_data['c_value'])
        ws_estimated_fx_summary.cell(row=row_number, column=4).value = to_excel_cell_value(row_data['d_value'])
        ws_estimated_fx_summary.cell(row=row_number, column=5).value = row_data['e_formula']
        ws_estimated_fx_summary.cell(row=row_number, column=6).value = row_data['f_formula']
        ws_estimated_fx_summary.cell(row=row_number, column=7).value = to_excel_cell_value(row_data['g_value'])
        ws_estimated_fx_summary.cell(row=row_number, column=8).value = to_excel_cell_value(row_data['h_value'])
        ws_estimated_fx_summary.cell(row=row_number, column=9).value = row_data['i_formula']
        ws_estimated_fx_summary.cell(row=row_number, column=10).value = row_data['j_formula']

    total_row_number = len(estimated_rows) + 2
    ws_estimated_fx_summary.cell(row=total_row_number, column=1).value = 'Grand Total'
    ws_estimated_fx_summary.cell(row=total_row_number, column=3).value = build_excel_formula(f"SUM(C2:C{total_row_number - 1})")
    ws_estimated_fx_summary.cell(row=total_row_number, column=6).value = build_excel_formula(f"SUM(F2:F{total_row_number - 1})")
    ws_estimated_fx_summary.cell(row=total_row_number, column=8).value = build_excel_formula(f"SUM(H2:H{total_row_number - 1})")
    ws_estimated_fx_summary.cell(row=total_row_number, column=9).value = build_excel_formula(f"SUM(I2:I{total_row_number - 1})")
    fill_fx_transaction_row(ws_estimated_fx_summary, total_row_number, 1, 10, FX_TRANSACTION_HEADER_FILL)
    apply_estimated_fx_summary_top_table_formats(ws_estimated_fx_summary, 2, total_row_number)

    validate_estimated_fx_summary_output(ws_estimated_fx_summary, estimated_rows)
    logging.info("Estimated FX Summary validated: wrote %s rows to 预估换汇汇总", len(estimated_rows))

    return {
        'estimated_fx_summary_rows': estimated_rows,
        'row_count': len(estimated_rows),
    }


#######################################################
#  数透<日期> / 预估换汇汇总 FX Transaction Analysis
#######################################################
def build_fx_transaction_sheet_name(transaction_dates):
    if normalize_cell_text(transaction_dates) == "":
        raise ValueError("transaction_dates is required to build the FX transaction analysis sheet name.")

    transformed_dates = transaction_dates.replace("yyyy-", "")
    transformed_dates = transformed_dates[5:] if len(transformed_dates) >= 5 and transformed_dates[4] == "-" else transformed_dates
    transformed_dates = transformed_dates.replace("-", "")
    return f"数透{transformed_dates}"


def recreate_fx_transaction_sheet(write_wb, transaction_dates):
    fx_transaction_sheet_name = build_fx_transaction_sheet_name(transaction_dates)
    if fx_transaction_sheet_name in write_wb.sheetnames:
        write_wb.remove(write_wb[fx_transaction_sheet_name])
    return write_wb.create_sheet(fx_transaction_sheet_name, 0)


def apply_fx_transaction_sheet_column_widths(ws_fx_transaction):
    ws_fx_transaction.column_dimensions['A'].width = 16
    ws_fx_transaction.column_dimensions['B'].width = 16
    ws_fx_transaction.column_dimensions['C'].width = 40
    ws_fx_transaction.column_dimensions['D'].width = 25
    ws_fx_transaction.column_dimensions['H'].width = 15
    ws_fx_transaction.column_dimensions['I'].width = 15
    ws_fx_transaction.column_dimensions['J'].width = 15
    ws_fx_transaction.column_dimensions['K'].width = 15


def write_fx_transaction_title_cell(ws, row_number, col_number, title):
    cell = ws.cell(row=row_number, column=col_number)
    cell.value = title
    cell.fill = FX_TRANSACTION_TITLE_FILL


def write_fx_transaction_header(ws, row_number, start_col, headers):
    for col_offset, header in enumerate(headers):
        cell = ws.cell(row=row_number, column=start_col + col_offset)
        cell.value = header
        cell.fill = FX_TRANSACTION_HEADER_FILL


def write_fx_transaction_rows(ws, start_row, start_col, rows):
    for row_offset, row_values in enumerate(rows):
        for col_offset, value in enumerate(row_values):
            ws.cell(row=start_row + row_offset, column=start_col + col_offset).value = to_excel_cell_value(value)


def fill_fx_transaction_row(ws, row_number, start_col, end_col, fill):
    for col_number in range(start_col, end_col + 1):
        ws.cell(row=row_number, column=col_number).fill = fill


def build_fx_transaction_table_a(estimated_rows):
    grouped = {}

    for row in estimated_rows:
        group_key = (normalize_cell_text(row['d_value']), normalize_cell_text(row['g_value']))
        bucket = grouped.setdefault(group_key, {'sum_f': 0.0, 'sum_i': 0.0})
        bucket['sum_f'] += numeric_cell_value(row['f_value'])
        bucket['sum_i'] += numeric_cell_value(row['i_value'])

    grouped_rows = []
    for group_key in sorted(grouped.keys()):
        totals = grouped[group_key]
        grouped_rows.append({
            'payout_currency': group_key[0],
            'settlement_currency': group_key[1],
            'sum_payout_amount': totals['sum_f'],
            'sum_settlement_amount': totals['sum_i'],
        })

    grand_total_row = {
        'payout_currency': 'Grand Total',
        'settlement_currency': None,
        'sum_payout_amount': sum(row['sum_payout_amount'] for row in grouped_rows),
        'sum_settlement_amount': sum(row['sum_settlement_amount'] for row in grouped_rows),
    }
    return grouped_rows, grand_total_row


def partition_fx_transaction_netting_rows(table_a_rows):
    row_lookup = {
        (row['payout_currency'], row['settlement_currency']): row
        for row in table_a_rows
    }
    visited_pairs = set()
    non_netting_rows = []
    netting_groups = []

    for row in table_a_rows:
        pair_key = (row['payout_currency'], row['settlement_currency'])
        if pair_key in visited_pairs:
            continue

        reverse_key = (pair_key[1], pair_key[0])
        reverse_row = row_lookup.get(reverse_key)
        if reverse_row is None or reverse_key == pair_key:
            non_netting_rows.append(row)
            visited_pairs.add(pair_key)
            continue

        row_one = dict(row)
        row_two = dict(reverse_row)
        if row_one['sum_payout_amount'] <= row_two['sum_settlement_amount']:
            row_one, row_two = row_two, row_one

        netting_groups.append({
            'row_one': row_one,
            'row_two': row_two,
        })
        visited_pairs.add(pair_key)
        visited_pairs.add(reverse_key)

    return non_netting_rows, netting_groups


def publish_fx_transaction_table_a(ws_fx_transaction, start_row, start_col, grouped_rows, grand_total_row):
    headers = [
        '打款币种',
        '清算币种',
        '求和项:预估通道打款金额（已扣除手续费3.2%）',
        '求和项:清算净额（扣除收费）',
    ]
    write_fx_transaction_title_cell(ws_fx_transaction, start_row, start_col, '表格A')
    write_fx_transaction_header(ws_fx_transaction, start_row + 1, start_col, headers)

    data_rows = [
        [row['payout_currency'], row['settlement_currency'], row['sum_payout_amount'], row['sum_settlement_amount']]
        for row in grouped_rows
    ]
    data_rows.append([
        grand_total_row['payout_currency'],
        grand_total_row['settlement_currency'],
        grand_total_row['sum_payout_amount'],
        grand_total_row['sum_settlement_amount'],
    ])
    write_fx_transaction_rows(ws_fx_transaction, start_row + 2, start_col, data_rows)
    fill_fx_transaction_row(ws_fx_transaction, start_row + 1, start_col, start_col + 3, FX_TRANSACTION_HEADER_FILL)
    fill_fx_transaction_row(
        ws_fx_transaction,
        start_row + 1 + len(data_rows),
        start_col,
        start_col + 3,
        FX_TRANSACTION_HEADER_FILL,
    )
    logging.info(
        "Published %s!表格A at A%s:D%s with %s grouped rows plus Grand Total",
        ws_fx_transaction.title,
        start_row,
        start_row + 1 + len(data_rows),
        len(grouped_rows),
    )
    return {
        'start_row': start_row,
        'end_row': start_row + 1 + len(data_rows),
        'rows': grouped_rows,
        'grand_total_row': grand_total_row,
    }


def publish_fx_transaction_table_b(ws_fx_transaction, start_row, start_col, non_netting_rows):
    headers = [
        '打款币种',
        '清算币种',
        '求和项:预估通道打款金额（已扣除手续费3.2%）',
        '求和项:清算净额（扣除收费）',
    ]
    write_fx_transaction_title_cell(ws_fx_transaction, start_row, start_col, '表格B(非轧差)')
    write_fx_transaction_header(ws_fx_transaction, start_row + 1, start_col, headers)
    write_fx_transaction_rows(
        ws_fx_transaction,
        start_row + 2,
        start_col,
        [[row['payout_currency'], row['settlement_currency'], row['sum_payout_amount'], row['sum_settlement_amount']] for row in non_netting_rows],
    )
    logging.info(
        "Published %s!表格B at A%s:D%s with %s non-netting rows",
        ws_fx_transaction.title,
        start_row,
        start_row + 1 + max(len(non_netting_rows), 1),
        len(non_netting_rows),
    )
    return {
        'start_row': start_row,
        'end_row': start_row + 1 + max(len(non_netting_rows), 1),
        'rows': non_netting_rows,
    }


def publish_fx_transaction_table_c(ws_fx_transaction, start_row, start_col, netting_groups):
    write_fx_transaction_title_cell(ws_fx_transaction, start_row, start_col, '表格C(轧差)')
    logging.info(
        "Publishing %s!表格C at H%s:K... for %s netting groups",
        ws_fx_transaction.title,
        start_row,
        len(netting_groups),
    )
    current_row = start_row + 1
    net_rows = []

    for index, group in enumerate(netting_groups):
        row_one_number = current_row
        write_fx_transaction_rows(
            ws_fx_transaction,
            row_one_number,
            start_col,
            [[
                group['row_one']['payout_currency'],
                group['row_one']['settlement_currency'],
                group['row_one']['sum_payout_amount'],
                group['row_one']['sum_settlement_amount'],
            ]],
        )
        current_row += 1

        row_two_number = current_row
        write_fx_transaction_rows(
            ws_fx_transaction,
            row_two_number,
            start_col,
            [[
                group['row_two']['payout_currency'],
                group['row_two']['settlement_currency'],
                group['row_two']['sum_payout_amount'],
                group['row_two']['sum_settlement_amount'],
            ]],
        )
        current_row += 1

        row_three_number = current_row
        row_three_col3_formula = build_excel_formula(
            f"{ws_fx_transaction.cell(row=row_one_number, column=start_col + 2).coordinate}-"
            f"{ws_fx_transaction.cell(row=row_two_number, column=start_col + 3).coordinate}"
        )
        row_three_col4_formula = build_excel_formula(
            f"{ws_fx_transaction.cell(row=row_one_number, column=start_col + 3).coordinate}-"
            f"{ws_fx_transaction.cell(row=row_two_number, column=start_col + 2).coordinate}"
        )
        write_fx_transaction_rows(
            ws_fx_transaction,
            row_three_number,
            start_col,
            [[
                group['row_one']['payout_currency'],
                group['row_one']['settlement_currency'],
                row_three_col3_formula,
                row_three_col4_formula,
            ]],
        )
        fill_fx_transaction_row(ws_fx_transaction, row_three_number, start_col, start_col + 3, FX_TRANSACTION_NET_FILL)
        net_rows.append({
            'payout_currency': group['row_one']['payout_currency'],
            'settlement_currency': group['row_one']['settlement_currency'],
            'sum_payout_amount': row_three_col3_formula,
            'sum_settlement_amount': row_three_col4_formula,
            'sum_payout_amount_value': group['row_one']['sum_payout_amount'] - group['row_two']['sum_settlement_amount'],
            'sum_settlement_amount_value': group['row_one']['sum_settlement_amount'] - group['row_two']['sum_payout_amount'],
            'is_net_result': True,
        })
        current_row += 1

        if index != len(netting_groups) - 1:
            current_row += 1

    end_row = max(start_row + 1, current_row - 1)
    logging.info(
        "Published %s!表格C at H%s:K%s with %s net-result rows",
        ws_fx_transaction.title,
        start_row,
        end_row,
        len(net_rows),
    )
    return {
        'start_row': start_row,
        'end_row': end_row,
        'rows': net_rows,
    }


def publish_fx_transaction_table_d(ws_fx_transaction, start_row, start_col, non_netting_rows, net_rows):
    headers = [
        '打款币种',
        '清算币种',
        '求和项:预估通道打款金额（已扣除手续费3.2%）',
        '求和项:清算净额（扣除收费）',
    ]
    combined_rows = [
        dict(
            row,
            is_net_result=False,
            sum_payout_amount_value=row['sum_payout_amount'],
            sum_settlement_amount_value=row['sum_settlement_amount'],
        )
        for row in non_netting_rows
    ] + [dict(row, is_net_result=True) for row in net_rows]
    combined_rows.sort(key=lambda row: (normalize_cell_text(row['payout_currency']), normalize_cell_text(row['settlement_currency'])))

    write_fx_transaction_title_cell(ws_fx_transaction, start_row, start_col, '表格D(将轧差后的数据一起汇总)')
    write_fx_transaction_header(ws_fx_transaction, start_row + 1, start_col, headers)
    current_row = start_row + 2
    for row_data in combined_rows:
        write_fx_transaction_rows(
            ws_fx_transaction,
            current_row,
            start_col,
            [[row_data['payout_currency'], row_data['settlement_currency'], row_data['sum_payout_amount'], row_data['sum_settlement_amount']]],
        )
        if row_data.get('is_net_result'):
            fill_fx_transaction_row(ws_fx_transaction, current_row, start_col, start_col + 3, FX_TRANSACTION_NET_RESULT_FILL)
        current_row += 1

    logging.info(
        "Published %s!表格D at A%s:D%s with %s combined rows (%s non-netting + %s net-result)",
        ws_fx_transaction.title,
        start_row,
        start_row + 1 + max(len(combined_rows), 1),
        len(combined_rows),
        len(non_netting_rows),
        len(net_rows),
    )
    return {
        'start_row': start_row,
        'end_row': start_row + 1 + max(len(combined_rows), 1),
        'rows': combined_rows,
    }


def publish_fx_transaction_table_e(ws_fx_transaction, start_row, start_col, table_d_rows):
    headers = ['卖出币种', '卖出金额', '买入币种', '买入金额']
    table_e_rows = [
        {
            'sell_currency': row['payout_currency'],
            'sell_amount': row['sum_payout_amount'],
            'sell_amount_value': row['sum_payout_amount_value'],
            'buy_currency': row['settlement_currency'],
            'buy_amount': row['sum_settlement_amount'],
            'buy_amount_value': row['sum_settlement_amount_value'],
        }
        for row in table_d_rows
    ]
    write_fx_transaction_title_cell(ws_fx_transaction, start_row, start_col, '表格E(盯盘所需数据)')
    write_fx_transaction_header(ws_fx_transaction, start_row + 1, start_col, headers)
    write_fx_transaction_rows(
        ws_fx_transaction,
        start_row + 2,
        start_col,
        [[row['sell_currency'], row['sell_amount'], row['buy_currency'], row['buy_amount']] for row in table_e_rows],
    )
    logging.info(
        "Published %s!表格E at A%s:D%s with %s rows for trading input",
        ws_fx_transaction.title,
        start_row,
        start_row + 1 + max(len(table_e_rows), 1),
        len(table_e_rows),
    )
    return {
        'start_row': start_row,
        'end_row': start_row + 1 + max(len(table_e_rows), 1),
        'rows': table_e_rows,
    }


def publish_fx_transaction_table_f(ws_fx_transaction, start_row, start_col, table_e_rows):
    headers = ['卖出币种', '卖出金额', '买入币种']
    write_fx_transaction_title_cell(ws_fx_transaction, start_row, start_col, '表格F(最终数据)')
    write_fx_transaction_header(ws_fx_transaction, start_row + 1, start_col, headers)
    logging.info(
        "Publishing %s!表格F at H%s:J... from %s 表格E rows with rounded amounts",
        ws_fx_transaction.title,
        start_row,
        len(table_e_rows),
    )
    table_f_rows = []
    for row_offset, row in enumerate(table_e_rows):
        table_e_row_number = start_row + 2 + row_offset
        rounded_formula = build_excel_formula(f"ROUND(B{table_e_row_number},-2)")
        rounded_value = round_to_nearest_hundred(row['sell_amount_value'])
        table_f_rows.append({
            'sell_currency': row['sell_currency'],
            'sell_amount': rounded_value,
            'sell_amount_formula': rounded_formula,
            'buy_currency': row['buy_currency'],
        })
        write_fx_transaction_rows(
            ws_fx_transaction,
            start_row + 2 + row_offset,
            start_col,
            [[row['sell_currency'], rounded_formula, row['buy_currency']]],
        )
    logging.info(
        "Published %s!表格F at H%s:J%s with %s final rows",
        ws_fx_transaction.title,
        start_row,
        start_row + 1 + max(len(table_f_rows), 1),
        len(table_f_rows),
    )
    return {
        'start_row': start_row,
        'end_row': start_row + 1 + max(len(table_f_rows), 1),
        'rows': table_f_rows,
    }


def build_fx_transaction_summary_remarks(summary_rows):
    def format_remark_amount(value):
        numeric_value = numeric_cell_value(value)
        if float(numeric_value).is_integer():
            return str(int(numeric_value))
        return str(numeric_value)

    remark_parts = []
    for row in summary_rows:
        sell_amount = numeric_cell_value(row['sell_amount'])
        if sell_amount == 0:
            continue
        if sell_amount > 0:
            remark_parts.append(
                f"{format_remark_amount(sell_amount)} {normalize_cell_text(row['sell_currency'])}换{normalize_cell_text(row['buy_currency'])}"
            )
        else:
            remark_parts.append(
                f"{normalize_cell_text(row['buy_currency'])}换{format_remark_amount(abs(sell_amount))} {normalize_cell_text(row['sell_currency'])}"
            )
    return '换汇时可以用' + '，'.join(remark_parts)


def append_fx_transaction_summary_table(write_wb, table_f_rows):
    ws_estimated_fx_summary = write_wb[ESTIMATED_FX_SUMMARY_SHEET_NAME]
    summary_header_row = get_last_data_row(ws_estimated_fx_summary, 1, 10, min_row=1) + 11
    logging.info(
        "Appending 预估换汇汇总 bottom summary table at A%s:E... based on %s 表格F rows",
        summary_header_row,
        len(table_f_rows),
    )
    summary_rows = [
        {
            'sell_currency': row['sell_currency'],
            'sell_amount': row['sell_amount'],
            'buy_currency': row['buy_currency'],
            'buy_amount': None,
        }
        for row in table_f_rows
    ]

    write_fx_transaction_header(
        ws_estimated_fx_summary,
        summary_header_row,
        1,
        ['卖出币种', '卖出金额', '买入币种', '买入金额', '备注'],
    )
    write_fx_transaction_rows(
        ws_estimated_fx_summary,
        summary_header_row + 1,
        1,
        [[row['sell_currency'], row['sell_amount'], row['buy_currency'], row['buy_amount']] for row in summary_rows],
    )
    ws_estimated_fx_summary.merge_cells(start_row=summary_header_row, start_column=7, end_row=summary_header_row, end_column=10)
    remarks_cell = ws_estimated_fx_summary.cell(row=summary_header_row, column=7)
    remarks_cell.value = build_fx_transaction_summary_remarks(summary_rows)
    remarks_cell.alignment = Alignment(wrap_text=True)

    ws_estimated_fx_summary.row_dimensions[summary_header_row].height = 50
    for col_number in range(1, 6):
        ws_estimated_fx_summary.cell(row=summary_header_row, column=col_number).alignment = Alignment(
            horizontal='center',
            vertical='center',
        )

    for row_number in range(summary_header_row + 1, summary_header_row + 1 + len(summary_rows)):
        ws_estimated_fx_summary.cell(row=row_number, column=1).alignment = Alignment(horizontal='center')
        ws_estimated_fx_summary.cell(row=row_number, column=2).number_format = ESTIMATED_FX_SUMMARY_NUMBER_2_FORMAT
        ws_estimated_fx_summary.cell(row=row_number, column=3).alignment = Alignment(horizontal='center')

    logging.info(
        "Appended 预估换汇汇总 bottom summary table at A%s:E%s and merged remarks G%s:J%s",
        summary_header_row,
        summary_header_row + len(summary_rows),
        summary_header_row,
        summary_header_row,
    )
    return {
        'summary_header_row': summary_header_row,
        'summary_rows': summary_rows,
    }


def publish_fx_transaction_analysis(write_wb, transaction_dates, estimated_rows):
    ws_fx_transaction = recreate_fx_transaction_sheet(write_wb, transaction_dates)
    logging.info(
        "Publishing FX transaction analysis sheet %s from %s estimated FX rows...",
        ws_fx_transaction.title,
        len(estimated_rows),
    )
    apply_fx_transaction_sheet_column_widths(ws_fx_transaction)
    table_a_rows, table_a_grand_total = build_fx_transaction_table_a(estimated_rows)
    non_netting_rows, netting_groups = partition_fx_transaction_netting_rows(table_a_rows)

    fx_transaction_table_a = publish_fx_transaction_table_a(ws_fx_transaction, 1, 1, table_a_rows, table_a_grand_total)
    fx_transaction_table_b = publish_fx_transaction_table_b(ws_fx_transaction, fx_transaction_table_a['end_row'] + 10, 1, non_netting_rows)
    fx_transaction_table_c = publish_fx_transaction_table_c(ws_fx_transaction, fx_transaction_table_b['start_row'], 8, netting_groups)
    fx_transaction_table_d = publish_fx_transaction_table_d(
        ws_fx_transaction,
        max(fx_transaction_table_b['end_row'], fx_transaction_table_c['end_row']) + 10,
        1,
        non_netting_rows,
        fx_transaction_table_c['rows'],
    )
    fx_transaction_table_e = publish_fx_transaction_table_e(ws_fx_transaction, fx_transaction_table_d['end_row'] + 10, 1, fx_transaction_table_d['rows'])
    fx_transaction_table_f = publish_fx_transaction_table_f(ws_fx_transaction, fx_transaction_table_e['start_row'], 8, fx_transaction_table_e['rows'])
    fx_transaction_summary_table = append_fx_transaction_summary_table(write_wb, fx_transaction_table_f['rows'])

    logging.info(
        "FX transaction analysis validated: created %s with A=%s rows, B=%s rows, C=%s net groups, D=%s rows, E=%s rows, F=%s rows",
        ws_fx_transaction.title,
        len(fx_transaction_table_a['rows']),
        len(fx_transaction_table_b['rows']),
        len(netting_groups),
        len(fx_transaction_table_d['rows']),
        len(fx_transaction_table_e['rows']),
        len(fx_transaction_table_f['rows']),
    )

    return {
        'fx_transaction_sheet_name': ws_fx_transaction.title,
        'fx_transaction_table_a': fx_transaction_table_a,
        'fx_transaction_table_b': fx_transaction_table_b,
        'fx_transaction_table_c': fx_transaction_table_c,
        'fx_transaction_table_d': fx_transaction_table_d,
        'fx_transaction_table_e': fx_transaction_table_e,
        'fx_transaction_table_f': fx_transaction_table_f,
        'fx_transaction_summary_table': fx_transaction_summary_table,
    }


#######################################################
#  Main Orchestration
#######################################################
def run_fx_consolidation_postprocess(workbook_path, log_path=None):
    workbook_path = os.path.abspath(workbook_path)
    if not os.path.isfile(workbook_path):
        raise FileNotFoundError(f"Input Excel file not found: {workbook_path}")

    if log_path is None:
        timestamp = datetime.now().strftime('%Y%m%d-%H%M%S')
        workbook_dir = os.path.dirname(workbook_path)
        log_path = os.path.join(workbook_dir, f'finalize_fx_summary_report_{timestamp}.log')

    configure_run_logging(log_path)
    logging.info("Starting FX consolidation post-processing...")
    logging.info("Input workbook: %s", workbook_path)

    # Load the completed workbook for value reads and formula-preserving writes.
    read_wb = load_workbook(workbook_path, data_only=True)
    write_wb = load_workbook(workbook_path)
    transaction_dates = None
    exchange_rate_lookup = {}
    estimated_fx_summary_results = {'estimated_fx_summary_rows': []}
    fx_transaction_results = {}

    try:
        if CHANNEL_ORDER_SHEET_NAME not in read_wb.sheetnames:
            raise KeyError(f"Source sheet not found in workbook: {CHANNEL_ORDER_SHEET_NAME}")
        if ACCOUNT_STATEMENT_SHEET_NAME not in read_wb.sheetnames:
            raise KeyError(f"Source sheet not found in workbook: {ACCOUNT_STATEMENT_SHEET_NAME}")
        if DAILY_EXCHANGE_RATE_SHEET_NAME not in read_wb.sheetnames:
            raise KeyError(f"Source sheet not found in workbook: {DAILY_EXCHANGE_RATE_SHEET_NAME}")

        ws_account_statement = read_wb[ACCOUNT_STATEMENT_SHEET_NAME]
        ws_daily_exchange_rate = read_wb[DAILY_EXCHANGE_RATE_SHEET_NAME]
        logging.info(
            "Loaded workbook sheets for finalization: %s rows in 账户流水, %s rows in 每日汇率(oc系统中获取）",
            max(ws_account_statement.max_row - 1, 0),
            max(ws_daily_exchange_rate.max_row - 1, 0),
        )

        # Rebuild 数据透视表 input rows from 渠道订单.
        ws_pivot, source_rows = rebuild_settlement_flow_input_sheet(write_wb, read_wb)

        # Build the grouped settlement-flow summary in 数据透视表.
        build_settlement_flow_summary(ws_pivot, source_rows)

        # Publish the final settlement-flow outputs to 1数透结果.
        publish_settlement_flow_results(write_wb, ws_pivot)

        # Generate 预估换汇汇总 and capture reusable computed values for later phases.
        try:
            transaction_dates = build_transaction_dates(ws_account_statement)
            exchange_rate_lookup = build_daily_exchange_rate_lookup(ws_daily_exchange_rate)
            estimated_fx_summary_rows = build_estimated_fx_summary_rows(
                write_wb[SETTLEMENT_FLOW_OUTPUT_SHEET_NAME],
                transaction_dates,
                exchange_rate_lookup,
            )
            estimated_fx_summary_results = publish_estimated_fx_summary(write_wb, estimated_fx_summary_rows)
        except Exception as exc:
            logging.error(
                "Skipping 预估换汇汇总 generation because of a downstream lookup/calculation error: %s",
                exc,
            )

        if transaction_dates and estimated_fx_summary_results['estimated_fx_summary_rows']:
            # Generate the FX transaction analysis sheet and append the summary table.
            fx_transaction_results = publish_fx_transaction_analysis(
                write_wb,
                transaction_dates,
                estimated_fx_summary_results['estimated_fx_summary_rows'],
            )
        else:
            logging.info(
                "Skipping FX transaction analysis because transaction_dates=%s and estimated row count=%s",
                transaction_dates,
                len(estimated_fx_summary_results['estimated_fx_summary_rows']),
            )

        if SUMMARY_SHEET_NAME in write_wb.sheetnames:
            logging.info("Removing intermediate sheet before final save: %s", SUMMARY_SHEET_NAME)
            del write_wb[SUMMARY_SHEET_NAME]

        # Save the updated workbook.
        logging.info("Saving finalized workbook updates back to: %s", workbook_path)
        write_wb.active = write_wb.sheetnames.index(ESTIMATED_FX_SUMMARY_SHEET_NAME)
        write_wb.save(workbook_path)
        logging.info("Completed FX consolidation post-processing: %s", workbook_path)
    finally:
        read_wb.close()
        write_wb.close()

    return {
        'workbook_path': workbook_path,
        'log_path': log_path,
        'transaction_dates': transaction_dates,
        'exchange_rate_lookup': exchange_rate_lookup,
        'estimated_fx_summary_rows': estimated_fx_summary_results['estimated_fx_summary_rows'],
        **fx_transaction_results,
    }


def finalize_fx_summary_report(workbook_path, log_path=None):
    result = run_fx_consolidation_postprocess(workbook_path, log_path=log_path)
    final_path = publish_final_report(result['workbook_path'])
    logging.info("Completed final FX summary report: %s", final_path)
    result['final_path'] = final_path
    result['workbook_path'] = final_path
    return result


def main():
    parser = argparse.ArgumentParser(description="Generate the final FX summary report from an in-progress workbook.")
    parser.add_argument("workbook", help="Path to the in-progress Excel workbook.")
    args = parser.parse_args()

    try:
        finalize_fx_summary_report(args.workbook)
    except Exception as exc:
        logging.error("FX summary finalization failed: %s", exc)
        raise


if __name__ == "__main__":
    main()
