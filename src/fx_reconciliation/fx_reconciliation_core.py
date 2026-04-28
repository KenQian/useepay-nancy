"""
FX Settlement Automation Tool
-----------------------------
Purpose:
    Automates the daily reconciliation of Foreign Exchange (FX) channel settlements.
    Processes account statements, validates channel orders against historical
    exceptions, and applies live Excel formulas for financial auditing.
"""

import os
import argparse
import shutil
import logging
import re
import time
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

# --- Configuration & Decoupling ---
FILE_CONFIG = {
    'type_1': {'pattern': r'^1(-.*)?\.xls$', 'desc': 'Refunds', 'header_keys': ['商户号', '订单类型']},
    'type_2': {'pattern': r'^2(-.*)?\.xls$', 'desc': 'Consumption', 'header_keys': ['商户号', '订单类型']},
    'type_3': {'pattern': r'^3(-.*)?\.xls$', 'desc': 'Channel Orders', 'header_keys': ['交易类型', '商户订单号']},
}

# Setup logging configuration
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%H:%M:%S'
)

READ_EXCEL_TEXT_KWARGS = {
    'dtype': str,
    'keep_default_na': False,
}
SUMMARY_SHEET_NAME = '处理摘要'
LOG_FORMAT = '%(asctime)s - %(levelname)s - %(message)s'
DATE_FORMAT = '%H:%M:%S'
SPECIAL_ORDER_DROP_PREFIXES = [
    'Delligent DE',
]
FX_RATE_SOURCE_FILENAME = '基本汇率.xlsx'
FX_RATE_TARGET_SHEET_NAME = '每日汇率(oc系统中获取）'
FX_RATE_SOURCE_SHEET_NAME = '基本汇率'
FX_RATE_HEADER_NAMES = {
    '汇率来源',
    '原币种',
    '目标币种',
    '现汇买入价',
    '现汇卖出价',
    '中间价',
}


def get_source_files(directory, pattern_str):
    """Finds all files in directory matching a regex pattern, sorted."""
    regex = re.compile(pattern_str, re.IGNORECASE)
    matched_files = [f for f in os.listdir(directory) if regex.match(f)]
    matched_files.sort()
    return [os.path.join(directory, f) for f in matched_files]


def file_has_header(path, header_keywords):
    """
    Detect whether the first row is a header row.

    We intentionally read with header=None first so stacked files always keep
    a consistent positional schema regardless of whether each source file has
    a header row.
    """
    preview_df = pd.read_excel(
        path,
        sheet_name=0,
        header=None,
        **READ_EXCEL_TEXT_KWARGS,
    )
    if preview_df.empty:
        return False

    first_row = preview_df.iloc[0].fillna('').astype(str)
    first_row_str = " ".join(first_row.tolist())
    return any(key in first_row_str for key in header_keywords)


def load_and_stack_files(file_paths, header_keywords):
    """Iteratively loads and stacks multiple files using a positional schema."""
    stacked_frames = []

    for path in file_paths:
        df = pd.read_excel(
            path,
            sheet_name=0,
            header=None,
            **READ_EXCEL_TEXT_KWARGS,
        )
        if df.empty:
            logging.info(f"Skipping empty file: {os.path.basename(path)}")
            continue

        has_header = file_has_header(path, header_keywords)
        if has_header:
            df = df.iloc[1:].reset_index(drop=True)

        # Normalize to a positional schema so mixed header/no-header source
        # files stack cleanly.
        df.columns = range(df.shape[1])

        # Drop fully blank rows that often appear at the tail of exported xls files.
        blank_row_mask = df.apply(
            lambda row: all(str(value).strip() == '' for value in row),
            axis=1,
        )
        df = df[~blank_row_mask].reset_index(drop=True)
        if df.empty:
            logging.info(f"Skipping file with only header/blank rows: {os.path.basename(path)}")
            continue

        logging.info(
            "Loaded %s rows from %s (header_detected=%s)",
            len(df),
            os.path.basename(path),
            has_header,
        )
        stacked_frames.append(df)

    if not stacked_frames:
        return pd.DataFrame()

    return pd.concat(stacked_frames, ignore_index=True)


def build_lookup_map(df, key_col_idx=0, value_col_idx=1):
    lookup = {}
    if df.empty:
        return lookup

    for _, row in df.iterrows():
        key = str(row.iloc[key_col_idx]).strip()
        value = str(row.iloc[value_col_idx]).strip()
        if key:
            lookup[key] = value

    return lookup


def find_last_non_empty_row(ws, column_index):
    for row_number in range(ws.max_row, 1, -1):
        cell_value = ws.cell(row=row_number, column=column_index).value
        if cell_value is None:
            continue
        if str(cell_value).strip() != "":
            return row_number
    return 1


def append_payout_currency_row(ws_payout, ap_value, ah_value, aj_value):
    insert_row = find_last_non_empty_row(ws_payout, 2) + 1

    ws_payout.cell(row=insert_row, column=2).value = ap_value
    ws_payout.cell(row=insert_row, column=4).value = ah_value
    ws_payout.cell(row=insert_row, column=5).value = aj_value
    ws_payout.cell(row=insert_row, column=7).value = f"=B{insert_row}&C{insert_row}&D{insert_row}&E{insert_row}"

    return insert_row


def append_a07_mapping_row(ws_a07, ah_value):
    insert_row = find_last_non_empty_row(ws_a07, 1) + 1
    ws_a07.cell(row=insert_row, column=1).value = ah_value
    return insert_row


def get_data_row_value(data_row, column_index):
    if 1 <= column_index <= len(data_row):
        value = data_row[column_index - 1]
        if value is None:
            return ""
        return str(value).strip()
    return ""


def should_drop_special_order(ah_value):
    return any(ah_value.startswith(prefix) for prefix in SPECIAL_ORDER_DROP_PREFIXES)


def resolve_aq_value(ap_value, ah_value, ab_value, a01_lookup, a07_lookup):
    if ap_value == "2号通道":
        return a01_lookup.get(ah_value, "")
    if ap_value == "A07":
        return f"{a07_lookup.get(ah_value, '')}{ah_value}" if ah_value else ""
    if ap_value == "7号通道":
        return ab_value
    return ""


def format_summary_table(title, headers, rows):
    if not rows:
        return f"{title}\n(no records)"

    normalized_rows = [[str(value) for value in row] for row in rows]
    widths = [len(header) for header in headers]
    for row in normalized_rows:
        for idx, value in enumerate(row):
            widths[idx] = max(widths[idx], len(value))

    def format_row(values):
        return " | ".join(value.ljust(widths[idx]) for idx, value in enumerate(values))

    separator = "-+-".join("-" * width for width in widths)
    table_lines = [
        title,
        format_row(headers),
        separator,
    ]
    table_lines.extend(format_row(row) for row in normalized_rows)
    return "\n".join(table_lines)


def to_excel_cell_value(value):
    if value is None:
        return None
    if isinstance(value, str) and value == "":
        return None
    return value


def get_sheet_header(ws, column_index):
    value = ws.cell(row=1, column=column_index).value
    if value is None:
        return f"Column {column_index}"
    value = str(value).strip()
    return value or f"Column {column_index}"


def get_latest_baseline(root_dir):
    """Finds the most recent non-WIP baseline file in the directory."""
    files = [f for f in os.listdir(root_dir) if f.startswith('各通道需换汇情况汇总-') and f.endswith('.xlsx') and '-wip' not in f]
    if not files:
        raise FileNotFoundError("Baseline file not found.")
    files.sort(reverse=True)
    return files[0]


def ensure_result_dir(root):
    result_dir = os.path.join(root, 'result')
    os.makedirs(result_dir, exist_ok=True)
    return result_dir


def configure_run_logging(log_path):
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)

    for handler in list(logger.handlers):
        if getattr(handler, '_fx_tool_file_handler', False):
            logger.removeHandler(handler)
            handler.close()

    file_handler = logging.FileHandler(log_path, encoding='utf-8')
    file_handler.setLevel(logging.INFO)
    file_handler.setFormatter(logging.Formatter(LOG_FORMAT, datefmt=DATE_FORMAT))
    file_handler._fx_tool_file_handler = True
    logger.addHandler(file_handler)


def discover_inputs(root):
    latest_baseline = get_latest_baseline(root)
    baseline_path = os.path.join(root, latest_baseline)
    fx_rate_path = os.path.join(root, FX_RATE_SOURCE_FILENAME)

    type1_files = get_source_files(root, FILE_CONFIG['type_1']['pattern'])
    type2_files = get_source_files(root, FILE_CONFIG['type_2']['pattern'])
    type3_files = get_source_files(root, FILE_CONFIG['type_3']['pattern'])

    logging.info(f"Baseline: {latest_baseline}")
    logging.info(
        "Found %s Refund files, %s Consumption files, %s Channel Order files",
        len(type1_files),
        len(type2_files),
        len(type3_files),
    )

    if not type1_files or not type2_files or not type3_files:
        raise FileNotFoundError("Missing one or more required source file types (1.xls, 2.xls, or 3.xls patterns).")
    if not os.path.isfile(fx_rate_path):
        raise FileNotFoundError(f"Required FX rate file not found: {FX_RATE_SOURCE_FILENAME}")

    return {
        'baseline_path': baseline_path,
        'fx_rate_path': fx_rate_path,
        'type1_files': type1_files,
        'type2_files': type2_files,
        'type3_files': type3_files,
    }


def create_wip_workbook(root, baseline_path):
    result_dir = ensure_result_dir(root)
    today_str = datetime.now().strftime('%Y%m%d')
    wip_filename = f"各通道需换汇情况汇总-{today_str}-wip.xlsx"
    wip_path = os.path.join(result_dir, wip_filename)
    shutil.copy2(baseline_path, wip_path)
    return wip_path


def load_source_data(type1_files, type2_files, type3_files):
    logging.info("Stacking %s Refund files for Module A...", len(type1_files))
    refunds_all = load_and_stack_files(type1_files, FILE_CONFIG['type_1']['header_keys'])
    logging.info("Stacking %s Consumption files for Module A...", len(type2_files))
    consumption_all = load_and_stack_files(type2_files, FILE_CONFIG['type_2']['header_keys'])

    account_statement_df = pd.concat([consumption_all, refunds_all], ignore_index=True)
    account_statement_df['Internal_Key_R'] = account_statement_df.iloc[:, 1].fillna('') + account_statement_df.iloc[:, 4].fillna('')

    logging.info("Stacking %s Channel Order files for Module B...", len(type3_files))
    channel_orders_raw = load_and_stack_files(type3_files, FILE_CONFIG['type_3']['header_keys'])
    channel_orders_filtered = channel_orders_raw[~channel_orders_raw.iloc[:, 9].isin(['预授权申请', '预授权撤销'])]

    return {
        'account_statement_df': account_statement_df,
        'channel_orders_filtered': channel_orders_filtered,
    }


def collect_revalidated_special_rows(baseline_path, account_statement_df):
    special_orders_df = pd.read_excel(
        baseline_path,
        sheet_name='特殊的渠道订单',
        **READ_EXCEL_TEXT_KWARGS,
    )
    valid_from_special = pd.DataFrame()
    revalidated_special_sheet_rows = []
    revalidated_special_rows = []

    if not special_orders_df.empty:
        special_orders_df['Worksheet_Row'] = special_orders_df.index + 2
        special_orders_df['Calc_AI_Key'] = special_orders_df.iloc[:, 3].fillna('') + special_orders_df.iloc[:, 4].fillna('')
        valid_from_special = special_orders_df[special_orders_df['Calc_AI_Key'].isin(account_statement_df['Internal_Key_R'])]
        for _, row in valid_from_special.iterrows():
            revalidated_special_sheet_rows.append(int(row['Worksheet_Row']))
            revalidated_special_rows.append([
                str(row.iloc[2]).strip(),
                str(row.iloc[0]).strip(),
                str(row.iloc[34]).strip(),
                str(row['Calc_AI_Key']).strip(),
            ])

    return {
        'valid_from_special': valid_from_special,
        'revalidated_special_sheet_rows': revalidated_special_sheet_rows,
        'revalidated_special_rows': revalidated_special_rows,
    }


def prepare_workbook_for_write(wip_path, revalidated_special_sheet_rows):
    wb = load_workbook(wip_path)

    ws_acc = wb['账户流水']
    ws_acc.delete_rows(2, ws_acc.max_row)

    ws_chan = wb['渠道订单']
    ws_chan.delete_rows(2, ws_chan.max_row)

    ws_spec = wb['特殊的渠道订单']
    for row_number in sorted(revalidated_special_sheet_rows, reverse=True):
        ws_spec.delete_rows(row_number, 1)

    return {
        'wb': wb,
        'ws_acc': ws_acc,
        'ws_chan': ws_chan,
        'ws_spec': ws_spec,
        'ws_payout': wb['打款币种'],
        'ws_a07': wb['二级商户号映射表-A07'],
    }


def resolve_fx_rate_source_sheet(wb):
    if FX_RATE_SOURCE_SHEET_NAME in wb.sheetnames:
        return wb[FX_RATE_SOURCE_SHEET_NAME]
    if len(wb.sheetnames) == 1:
        return wb[wb.sheetnames[0]]
    raise ValueError(
        f"{FX_RATE_SOURCE_FILENAME} must contain sheet '{FX_RATE_SOURCE_SHEET_NAME}' "
        "or have exactly one sheet."
    )


def fx_rate_source_has_header(ws):
    first_row_values = []
    for row in ws.iter_rows(min_row=1, max_row=1, min_col=1, max_col=8, values_only=True):
        first_row_values = [str(value).strip() for value in row if value is not None]
    return any(value in FX_RATE_HEADER_NAMES for value in first_row_values)


def refresh_daily_fx_rate_sheet(wb, fx_rate_path):
    if FX_RATE_TARGET_SHEET_NAME not in wb.sheetnames:
        raise KeyError(f"Target sheet not found in workbook: {FX_RATE_TARGET_SHEET_NAME}")

    target_ws = wb[FX_RATE_TARGET_SHEET_NAME]
    if target_ws.max_row > 1:
        target_ws.delete_rows(2, target_ws.max_row - 1)

    source_wb = load_workbook(fx_rate_path, data_only=False)
    try:
        source_ws = resolve_fx_rate_source_sheet(source_wb)
        has_header = fx_rate_source_has_header(source_ws)
        start_row = 2 if has_header else 1
        target_row = 2
        imported_rows = 0

        logging.info(
            "Resolved FX rate source sheet: workbook=%s sheet=%s",
            os.path.basename(fx_rate_path),
            source_ws.title,
        )

        for row in source_ws.iter_rows(min_row=start_row, min_col=1, max_col=8, values_only=True):
            if all(value is None or str(value).strip() == "" for value in row):
                continue
            for col_idx, value in enumerate(row, start=1):
                target_ws.cell(row=target_row, column=col_idx).value = value
            target_ws.cell(row=target_row, column=9).value = f"=D{target_row}&E{target_row}"
            target_row += 1
            imported_rows += 1

        logging.info(
            "Refreshed %s with %s FX rate rows from %s",
            FX_RATE_TARGET_SHEET_NAME,
            imported_rows,
            source_ws.title,
        )
    finally:
        source_wb.close()


def build_target_channel_data(channel_orders_filtered, valid_from_special):
    target_data = []

    for _, row in channel_orders_filtered.iterrows():
        new_row = list(row.iloc[0:4]) + list(row.iloc[5:34]) + [row.iloc[4]]
        target_data.append(new_row)

    if not valid_from_special.empty:
        for _, row in valid_from_special.iterrows():
            target_data.append(list(row.iloc[0:34]))

    return target_data


def load_mapping_context(wip_path):
    mapping_sheets = {
        '打款币种': pd.read_excel(wip_path, sheet_name='打款币种', **READ_EXCEL_TEXT_KWARGS),
        '渠道名称': pd.read_excel(wip_path, sheet_name='渠道名称', **READ_EXCEL_TEXT_KWARGS),
        '二级商户号映射表-A01': pd.read_excel(wip_path, sheet_name='二级商户号映射表-A01', **READ_EXCEL_TEXT_KWARGS),
        '二级商户号映射表-A07': pd.read_excel(wip_path, sheet_name='二级商户号映射表-A07', **READ_EXCEL_TEXT_KWARGS),
    }

    return {
        'payout_lookup': build_lookup_map(mapping_sheets['打款币种'], key_col_idx=6, value_col_idx=5),
        'chan_map': build_lookup_map(mapping_sheets['渠道名称'], key_col_idx=0, value_col_idx=1),
        'a01_lookup': build_lookup_map(mapping_sheets['二级商户号映射表-A01'], key_col_idx=0, value_col_idx=1),
        'a07_lookup': build_lookup_map(mapping_sheets['二级商户号映射表-A07'], key_col_idx=0, value_col_idx=2),
        'pending_payout_keys': set(),
        'pending_a07_keys': set(),
        'payout_rows_added': [],
        'a07_rows_added': [],
        'special_rows_added': [],
    }


def append_special_order_row(ws_spec, data_row, next_spec):
    for c_idx, val in enumerate(data_row, start=1):
        ws_spec.cell(row=next_spec, column=c_idx).value = to_excel_cell_value(val)
    ws_spec.cell(row=next_spec, column=35).value = f"=D{next_spec}&E{next_spec}"
    ws_spec.cell(row=next_spec, column=36).value = f"=XLOOKUP(A{next_spec},账户流水!$R:$R,账户流水!$R:$R)"
    ws_spec.cell(row=next_spec, column=37).value = f"=LEFTB(Y{next_spec},10)"


def write_channel_order_row(ws_chan, curr_row, data_row):
    for c_idx, val in enumerate(data_row, start=1):
        ws_chan.cell(row=curr_row, column=c_idx).value = to_excel_cell_value(val)

    ws_chan.cell(row=curr_row, column=35).value = f"=D{curr_row}&E{curr_row}"
    ws_chan.cell(row=curr_row, column=36).value = f"=N{curr_row}"
    ws_chan.cell(row=curr_row, column=37).value = f'=IF(I{curr_row}="退款", -M{curr_row}*(1-0.032), M{curr_row}*(1-0.032))'
    ws_chan.cell(row=curr_row, column=38).value = f"=XLOOKUP(AR{curr_row}, 打款币种!$G:$G, 打款币种!$F:$F)"
    ws_chan.cell(row=curr_row, column=39).value = f"=VLOOKUP(AI{curr_row}, 账户流水!$R:$T, 2, 0)"
    ws_chan.cell(row=curr_row, column=40).value = f"=VLOOKUP(AI{curr_row}, 账户流水!$R:$T, 3, 0)"
    ws_chan.cell(row=curr_row, column=41).value = f"=IF(AL{curr_row}=AM{curr_row}, \"是\", \"否\")"
    ws_chan.cell(row=curr_row, column=42).value = f"=VLOOKUP(F{curr_row}, 渠道名称!$A:$B, 2, 0)"

    aq_f = (f'IF(AP{curr_row}="2号通道", XLOOKUP(AH{curr_row}, \'二级商户号映射表-A01\'!$A:$A, \'二级商户号映射表-A01\'!$B:$B), '
            f'IF(AP{curr_row}="A07", XLOOKUP(AH{curr_row}, \'二级商户号映射表-A07\'!$A:$A, \'二级商户号映射表-A07\'!$C:$C) & AH{curr_row}, '
            f'IF(AP{curr_row}="7号通道", AB{curr_row}, "")))')
    ws_chan.cell(row=curr_row, column=43).value = f"={aq_f}"
    ws_chan.cell(row=curr_row, column=44).value = f"=AP{curr_row}&AQ{curr_row}&AJ{curr_row}"


def process_target_channel_data(target_data, account_statement_df, worksheet_handles, mapping_context):
    keys_to_remove_from_statement = set()
    curr_row = 2
    exceptions_moved = 0
    dropped_special_rows_count = 0
    loop_start = time.perf_counter()
    account_statement_keys = set(account_statement_df['Internal_Key_R'])

    ws_chan = worksheet_handles['ws_chan']
    ws_spec = worksheet_handles['ws_spec']
    ws_payout = worksheet_handles['ws_payout']
    ws_a07 = worksheet_handles['ws_a07']

    total_rows = len(target_data)
    for idx, data_row in enumerate(target_data, start=1):
        if idx == 1 or idx % 5000 == 0 or idx == total_rows:
            logging.info(
                "Processing 渠道订单 rows: %s/%s elapsed=%.1fs",
                idx,
                total_rows,
                time.perf_counter() - loop_start,
            )
        col_f, col_d, col_e = data_row[5], data_row[3], data_row[4]
        ai_key = f"{col_d}{col_e}"
        ap_val = str(mapping_context['chan_map'].get(col_f, "")).strip()
        is_na = ai_key not in account_statement_keys

        if is_na and ap_val.lower() not in ["paypal", "afterpay直连"]:
            ah_value = get_data_row_value(data_row, 34)
            if should_drop_special_order(ah_value):
                keys_to_remove_from_statement.add(ai_key)
                exceptions_moved += 1
                dropped_special_rows_count += 1
                continue

            next_spec = ws_spec.max_row + 1
            append_special_order_row(ws_spec, data_row, next_spec)
            mapping_context['special_rows_added'].append([
                next_spec,
                get_data_row_value(data_row, 3),
                get_data_row_value(data_row, 1),
                ap_val,
                ai_key,
            ])
            keys_to_remove_from_statement.add(ai_key)
            exceptions_moved += 1
            continue

        ah_value = get_data_row_value(data_row, 34)
        aj_value = get_data_row_value(data_row, 14)
        ab_value = get_data_row_value(data_row, 28)

        if (
            ap_val == "A07"
            and ah_value
            and ah_value not in mapping_context['a07_lookup']
            and ah_value not in mapping_context['pending_a07_keys']
        ):
            insert_row = append_a07_mapping_row(ws_a07, ah_value)
            mapping_context['a07_lookup'][ah_value] = ""
            mapping_context['pending_a07_keys'].add(ah_value)
            mapping_context['a07_rows_added'].append([insert_row, ah_value])

        aq_value = resolve_aq_value(
            ap_val,
            ah_value,
            ab_value,
            mapping_context['a01_lookup'],
            mapping_context['a07_lookup'],
        )
        ar_value = f"{ap_val}{aq_value}{aj_value}"

        if (
            ap_val
            and aq_value
            and aj_value
            and ar_value not in mapping_context['payout_lookup']
            and ar_value not in mapping_context['pending_payout_keys']
        ):
            insert_row = append_payout_currency_row(ws_payout, ap_val, ah_value, aj_value)
            mapping_context['payout_lookup'][ar_value] = ""
            mapping_context['pending_payout_keys'].add(ar_value)
            mapping_context['payout_rows_added'].append([insert_row, ap_val, ah_value, aj_value, ar_value])

        write_channel_order_row(ws_chan, curr_row, data_row)
        curr_row += 1

    return {
        'keys_to_remove_from_statement': keys_to_remove_from_statement,
        'exceptions_moved': exceptions_moved,
        'dropped_special_rows_count': dropped_special_rows_count,
        'payout_rows_added': mapping_context['payout_rows_added'],
        'a07_rows_added': mapping_context['a07_rows_added'],
        'special_rows_added': mapping_context['special_rows_added'],
    }


def write_account_statement_sheet(ws_acc, account_statement_df):
    loop_start = time.perf_counter()
    total_rows = len(account_statement_df.index)
    for idx, row in enumerate(account_statement_df.values, start=1):
        if idx == 1 or idx % 5000 == 0 or idx == total_rows:
            logging.info(
                "Writing 账户流水 rows: %s/%s elapsed=%.1fs",
                idx,
                total_rows,
                time.perf_counter() - loop_start,
            )
        r_idx = idx + 1
        if pd.notna(row[0]) and str(row[0]).strip() != "":
            for c_idx, value in enumerate(row[:-1], start=1):
                ws_acc.cell(row=r_idx, column=c_idx).value = to_excel_cell_value(value)

            ws_acc.cell(row=r_idx, column=13).value = f"=K{r_idx}-L{r_idx}"
            ws_acc.cell(row=r_idx, column=18).value = f"=B{r_idx}&E{r_idx}"
            ws_acc.cell(row=r_idx, column=19).value = f"=G{r_idx}"
            ws_acc.cell(row=r_idx, column=20).value = f"=M{r_idx}"


def log_summary_tables(revalidated_special_rows, dropped_special_rows_count, a07_rows_added, payout_rows_added, special_rows_added, ws_a07, ws_payout):
    logging.info("Dropped rows while moving to 特殊的渠道订单 (AH starts with Delligent DE): %s", dropped_special_rows_count)
    logging.info(
        "\n%s",
        format_summary_table(
            "Added back to 渠道订单 from 特殊的渠道订单",
            ["C-交易流水号", "A-渠道订单号", "AP-通道名称", "AI-Key"],
            revalidated_special_rows,
        ),
    )
    logging.info(
        "\n%s",
        format_summary_table(
            "Added rows in 二级商户号映射表-A07",
            [
                "Row",
                f"A-{get_sheet_header(ws_a07, 1)}",
            ],
            a07_rows_added,
        ),
    )
    logging.info(
        "\n%s",
        format_summary_table(
            "Added rows in 打款币种",
            [
                "Row",
                f"B-{get_sheet_header(ws_payout, 2)}",
                f"D-{get_sheet_header(ws_payout, 4)}",
                f"E-{get_sheet_header(ws_payout, 5)}",
                f"G-{get_sheet_header(ws_payout, 7)}",
            ],
            payout_rows_added,
        ),
    )
    logging.info(
        "\n%s",
        format_summary_table(
            "Added rows in 特殊的渠道订单",
            ["Row", "C-交易流水号", "A-渠道订单号", "AP-通道名称", "AI-Key"],
            special_rows_added,
        ),
    )


def write_summary_sheet(wb, source_root, final_path, log_path, revalidated_special_rows, dropped_special_rows_count, a07_rows_added, payout_rows_added, special_rows_added):
    if SUMMARY_SHEET_NAME in wb.sheetnames:
        ws_summary = wb[SUMMARY_SHEET_NAME]
        ws_summary.delete_rows(1, ws_summary.max_row)
    else:
        channel_order_index = wb.sheetnames.index('渠道订单')
        ws_summary = wb.create_sheet(SUMMARY_SHEET_NAME, channel_order_index)

    summary_tables = [
        (
            "Added back to 渠道订单 from 特殊的渠道订单",
            ["C-交易流水号", "A-渠道订单号", "AP-通道名称", "AI-Key"],
            revalidated_special_rows,
        ),
        (
            "Added rows in 二级商户号映射表-A07",
            ["Row", "A-二级商户号"],
            a07_rows_added,
        ),
        (
            "Added rows in 打款币种",
            ["Row", "B-通道名称", "D-二级商户号", "E-交易币种", "G-Column 7"],
            payout_rows_added,
        ),
        (
            "Added rows in 特殊的渠道订单",
            ["Row", "C-交易流水号", "A-渠道订单号", "AP-通道名称", "AI-Key"],
            special_rows_added,
        ),
    ]

    current_row = 1
    ws_summary.cell(row=current_row, column=1).value = "处理摘要"
    current_row += 2

    metadata_rows = [
        ("运行时间", datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
        ("源文件夹", source_root),
        ("输出文件", final_path),
        ("日志文件", log_path),
        ("回补到渠道订单", len(revalidated_special_rows)),
        ("丢弃的 Delligent DE 特殊订单", dropped_special_rows_count),
        ("新增二级商户号映射表-A07", len(a07_rows_added)),
        ("新增打款币种", len(payout_rows_added)),
        ("新增特殊的渠道订单", len(special_rows_added)),
    ]

    for label, value in metadata_rows:
        ws_summary.cell(row=current_row, column=1).value = label
        ws_summary.cell(row=current_row, column=2).value = value
        current_row += 1

    current_row += 1

    for title, headers, rows in summary_tables:
        ws_summary.cell(row=current_row, column=1).value = title
        current_row += 1

        for col_idx, header in enumerate(headers, start=1):
            ws_summary.cell(row=current_row, column=col_idx).value = header
        current_row += 1

        if rows:
            for row in rows:
                for col_idx, value in enumerate(row, start=1):
                    ws_summary.cell(row=current_row, column=col_idx).value = value
                current_row += 1
        else:
            ws_summary.cell(row=current_row, column=1).value = "(no records)"
            current_row += 1

        current_row += 1


def finalize_workbook(wb, wip_path):
    save_start = time.perf_counter()
    logging.info("Saving workbook to WIP path...")
    final_path = wip_path.replace("-wip.xlsx", ".xlsx")
    wb.save(wip_path)
    logging.info("Workbook save completed in %.1fs", time.perf_counter() - save_start)
    rename_start = time.perf_counter()
    if os.path.exists(final_path):
        os.remove(final_path)
    os.rename(wip_path, final_path)
    logging.info("Workbook rename completed in %.1fs", time.perf_counter() - rename_start)
    logging.info(f"COMPLETED. File: {final_path}")
    return final_path


def run_fx_reconciliation(root, log_path=None):
    root = os.path.abspath(root)
    if not os.path.isdir(root):
        raise FileNotFoundError(f"Source folder not found: {root}")

    result_dir = ensure_result_dir(root)
    if log_path is None:
        timestamp = datetime.now().strftime('%Y%m%d-%H%M%S')
        log_path = os.path.join(result_dir, f'fx_reconciliation_{timestamp}.log')

    configure_run_logging(log_path)
    logging.info("Starting FX Settlement Automation...")
    run_start = time.perf_counter()

    phase_start = time.perf_counter()
    input_paths = discover_inputs(root)
    logging.info("Phase complete: discover_inputs elapsed=%.1fs", time.perf_counter() - phase_start)

    phase_start = time.perf_counter()
    wip_path = create_wip_workbook(root, input_paths['baseline_path'])
    logging.info("Phase complete: create_wip_workbook elapsed=%.1fs", time.perf_counter() - phase_start)

    phase_start = time.perf_counter()
    source_data = load_source_data(
        input_paths['type1_files'],
        input_paths['type2_files'],
        input_paths['type3_files'],
    )
    account_statement_df = source_data['account_statement_df']
    channel_orders_filtered = source_data['channel_orders_filtered']
    logging.info("Phase complete: load_source_data elapsed=%.1fs", time.perf_counter() - phase_start)

    phase_start = time.perf_counter()
    special_row_data = collect_revalidated_special_rows(input_paths['baseline_path'], account_statement_df)
    valid_from_special = special_row_data['valid_from_special']
    revalidated_special_rows = special_row_data['revalidated_special_rows']
    logging.info("Phase complete: collect_revalidated_special_rows elapsed=%.1fs", time.perf_counter() - phase_start)

    phase_start = time.perf_counter()
    workbook_state = prepare_workbook_for_write(
        wip_path,
        special_row_data['revalidated_special_sheet_rows'],
    )
    wb = workbook_state['wb']
    ws_acc = workbook_state['ws_acc']
    logging.info("Phase complete: prepare_workbook_for_write elapsed=%.1fs", time.perf_counter() - phase_start)

    phase_start = time.perf_counter()
    refresh_daily_fx_rate_sheet(wb, input_paths['fx_rate_path'])
    logging.info("Phase complete: refresh_daily_fx_rate_sheet elapsed=%.1fs", time.perf_counter() - phase_start)

    phase_start = time.perf_counter()
    target_data = build_target_channel_data(channel_orders_filtered, valid_from_special)
    logging.info("Phase complete: build_target_channel_data elapsed=%.1fs", time.perf_counter() - phase_start)

    phase_start = time.perf_counter()
    mapping_context = load_mapping_context(wip_path)
    logging.info("Phase complete: load_mapping_context elapsed=%.1fs", time.perf_counter() - phase_start)

    phase_start = time.perf_counter()
    processing_results = process_target_channel_data(
        target_data,
        account_statement_df,
        workbook_state,
        mapping_context,
    )
    logging.info("Phase complete: process_target_channel_data elapsed=%.1fs", time.perf_counter() - phase_start)

    if processing_results['keys_to_remove_from_statement']:
        account_statement_df = account_statement_df[
            ~account_statement_df['Internal_Key_R'].isin(processing_results['keys_to_remove_from_statement'])
        ]

    phase_start = time.perf_counter()
    write_account_statement_sheet(ws_acc, account_statement_df)
    logging.info("Phase complete: write_account_statement_sheet elapsed=%.1fs", time.perf_counter() - phase_start)

    logging.info(f"Processing finished. Exceptions: {processing_results['exceptions_moved']}")
    phase_start = time.perf_counter()
    log_summary_tables(
        revalidated_special_rows,
        processing_results['dropped_special_rows_count'],
        processing_results['a07_rows_added'],
        processing_results['payout_rows_added'],
        processing_results['special_rows_added'],
        workbook_state['ws_a07'],
        workbook_state['ws_payout'],
    )
    logging.info("Phase complete: log_summary_tables elapsed=%.1fs", time.perf_counter() - phase_start)

    phase_start = time.perf_counter()
    write_summary_sheet(
        wb,
        root,
        wip_path.replace("-wip.xlsx", ".xlsx"),
        log_path,
        revalidated_special_rows,
        processing_results['dropped_special_rows_count'],
        processing_results['a07_rows_added'],
        processing_results['payout_rows_added'],
        processing_results['special_rows_added'],
    )
    logging.info("Phase complete: write_summary_sheet elapsed=%.1fs", time.perf_counter() - phase_start)

    phase_start = time.perf_counter()
    final_path = finalize_workbook(wb, wip_path)
    logging.info("Phase complete: finalize_workbook elapsed=%.1fs", time.perf_counter() - phase_start)
    logging.info("Run completed in %.1fs", time.perf_counter() - run_start)
    return {
        'final_path': final_path,
        'log_path': log_path,
        'revalidated_special_rows': revalidated_special_rows,
        'a07_rows_added': processing_results['a07_rows_added'],
        'payout_rows_added': processing_results['payout_rows_added'],
        'special_rows_added': processing_results['special_rows_added'],
    }


def main():
    parser = argparse.ArgumentParser(description="FX Channel Settlement Automation Tool")
    parser.add_argument("directory", help="The root directory containing source files.")

    import sys
    if len(sys.argv) == 1:
        parser.print_help(sys.stderr); sys.exit(1)

    args = parser.parse_args()
    root = args.directory
    try:
        run_fx_reconciliation(root)
    except Exception as e:
        logging.error(f"Initialization Failed: {e}")
        return


if __name__ == "__main__":
    main()
