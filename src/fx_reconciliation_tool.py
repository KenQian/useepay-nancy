"""
FX Settlement Automation Tool
-----------------------------
Purpose:
    Automates the daily reconciliation of Foreign Exchange (FX) channel settlements.
    Processes account statements, validates channel orders against historical
    exceptions, and applies live Excel formulas for financial auditing.
"""

import os
import argparse
import shutil
import logging
import re
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

# --- Configuration & Decoupling ---
FILE_CONFIG = {
    'type_1': {'pattern': r'^1(-.*)?\.xls$', 'desc': 'Refunds', 'header_keys': ['商户号', '订单类型']},
    'type_2': {'pattern': r'^2(-.*)?\.xls$', 'desc': 'Consumption', 'header_keys': ['商户号', '订单类型']},
    'type_3': {'pattern': r'^3(-.*)?\.xls$', 'desc': 'Channel Orders', 'header_keys': ['交易类型', '商户订单号']},
}

# Setup logging configuration
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%H:%M:%S'
)

READ_EXCEL_TEXT_KWARGS = {
    'dtype': str,
    'keep_default_na': False,
}


def get_source_files(directory, pattern_str):
    """Finds all files in directory matching a regex pattern, sorted."""
    regex = re.compile(pattern_str, re.IGNORECASE)
    matched_files = [f for f in os.listdir(directory) if regex.match(f)]
    matched_files.sort()
    return [os.path.join(directory, f) for f in matched_files]


def file_has_header(path, header_keywords):
    """
    Detect whether the first row is a header row.

    We intentionally read with header=None first so stacked files always keep
    a consistent positional schema regardless of whether each source file has
    a header row.
    """
    preview_df = pd.read_excel(
        path,
        sheet_name=0,
        header=None,
        **READ_EXCEL_TEXT_KWARGS,
    )
    if preview_df.empty:
        return False

    first_row = preview_df.iloc[0].fillna('').astype(str)
    first_row_str = " ".join(first_row.tolist())
    return any(key in first_row_str for key in header_keywords)


def load_and_stack_files(file_paths, header_keywords):
    """Iteratively loads and stacks multiple files using a positional schema."""
    stacked_frames = []

    for path in file_paths:
        df = pd.read_excel(
            path,
            sheet_name=0,
            header=None,
            **READ_EXCEL_TEXT_KWARGS,
        )
        if df.empty:
            logging.info(f"Skipping empty file: {os.path.basename(path)}")
            continue

        has_header = file_has_header(path, header_keywords)
        if has_header:
            df = df.iloc[1:].reset_index(drop=True)

        # Normalize to a positional schema so mixed header/no-header source
        # files stack cleanly.
        df.columns = range(df.shape[1])

        # Drop fully blank rows that often appear at the tail of exported xls files.
        blank_row_mask = df.apply(
            lambda row: all(str(value).strip() == '' for value in row),
            axis=1,
        )
        df = df[~blank_row_mask].reset_index(drop=True)
        if df.empty:
            logging.info(f"Skipping file with only header/blank rows: {os.path.basename(path)}")
            continue

        logging.info(
            "Loaded %s rows from %s (header_detected=%s)",
            len(df),
            os.path.basename(path),
            has_header,
        )
        stacked_frames.append(df)

    if not stacked_frames:
        return pd.DataFrame()

    return pd.concat(stacked_frames, ignore_index=True)


def build_lookup_map(df, key_col_idx=0, value_col_idx=1):
    lookup = {}
    if df.empty:
        return lookup

    for _, row in df.iterrows():
        key = str(row.iloc[key_col_idx]).strip()
        value = str(row.iloc[value_col_idx]).strip()
        if key:
            lookup[key] = value

    return lookup


def find_last_non_empty_row(ws, column_index):
    for row_number in range(ws.max_row, 1, -1):
        cell_value = ws.cell(row=row_number, column=column_index).value
        if cell_value is None:
            continue
        if str(cell_value).strip() != "":
            return row_number
    return 1


def append_payout_currency_row(ws_payout, ap_value, ah_value, aj_value):
    insert_row = find_last_non_empty_row(ws_payout, 2) + 1

    ws_payout.cell(row=insert_row, column=2).value = ap_value
    ws_payout.cell(row=insert_row, column=4).value = ah_value
    ws_payout.cell(row=insert_row, column=5).value = aj_value
    ws_payout.cell(row=insert_row, column=7).value = f"=B{insert_row}&C{insert_row}&D{insert_row}&E{insert_row}"

    return insert_row


def append_a07_mapping_row(ws_a07, ah_value):
    insert_row = find_last_non_empty_row(ws_a07, 1) + 1
    ws_a07.cell(row=insert_row, column=1).value = ah_value
    return insert_row


def get_data_row_value(data_row, column_index):
    if 1 <= column_index <= len(data_row):
        value = data_row[column_index - 1]
        if value is None:
            return ""
        return str(value).strip()
    return ""


def resolve_aq_value(ap_value, ah_value, ab_value, a01_lookup, a07_lookup):
    if ap_value == "2号通道":
        return a01_lookup.get(ah_value, "")
    if ap_value == "A07":
        return f"{a07_lookup.get(ah_value, '')}{ah_value}" if ah_value else ""
    if ap_value == "7号通道":
        return ab_value
    return ""


def format_summary_table(title, headers, rows):
    if not rows:
        return f"{title}\n(no records)"

    normalized_rows = [[str(value) for value in row] for row in rows]
    widths = [len(header) for header in headers]
    for row in normalized_rows:
        for idx, value in enumerate(row):
            widths[idx] = max(widths[idx], len(value))

    def format_row(values):
        return " | ".join(value.ljust(widths[idx]) for idx, value in enumerate(values))

    separator = "-+-".join("-" * width for width in widths)
    table_lines = [
        title,
        format_row(headers),
        separator,
    ]
    table_lines.extend(format_row(row) for row in normalized_rows)
    return "\n".join(table_lines)


def to_excel_cell_value(value):
    if value is None:
        return None
    if isinstance(value, str) and value == "":
        return None
    return value


def get_sheet_header(ws, column_index):
    value = ws.cell(row=1, column=column_index).value
    if value is None:
        return f"Column {column_index}"
    value = str(value).strip()
    return value or f"Column {column_index}"


def get_latest_baseline(root_dir):
    """Finds the most recent non-WIP baseline file in the directory."""
    files = [f for f in os.listdir(root_dir) if f.startswith('各通道需换汇情况汇总-') and f.endswith('.xlsx') and '-wip' not in f]
    if not files:
        raise FileNotFoundError("Baseline file not found.")
    files.sort(reverse=True)
    return files[0]


def main():
    parser = argparse.ArgumentParser(description="FX Channel Settlement Automation Tool")
    parser.add_argument("directory", help="The root directory containing source files.")

    import sys
    if len(sys.argv) == 1:
        parser.print_help(sys.stderr); sys.exit(1)

    args = parser.parse_args()
    root = args.directory

    logging.info("Starting FX Settlement Automation...")

    # 1. File Discovery
    try:
        latest_baseline = get_latest_baseline(root)
        baseline_path = os.path.join(root, latest_baseline)

        type1_files = get_source_files(root, FILE_CONFIG['type_1']['pattern'])
        type2_files = get_source_files(root, FILE_CONFIG['type_2']['pattern'])
        type3_files = get_source_files(root, FILE_CONFIG['type_3']['pattern'])

        logging.info(f"Baseline: {latest_baseline}")
        logging.info(
            "Found %s Refund files, %s Consumption files, %s Channel Order files",
            len(type1_files),
            len(type2_files),
            len(type3_files),
        )

        if not type1_files or not type2_files or not type3_files:
            logging.error("Missing one or more required source file types (1.xls, 2.xls, or 3.xls patterns).")
            return

    except Exception as e:
        logging.error(f"Initialization Failed: {e}"); return

    # 2. Workspace Creation
    today_str = datetime.now().strftime('%Y%m%d')
    wip_filename = f"各通道需换汇情况汇总-{today_str}-wip.xlsx"
    wip_path = os.path.join(root, wip_filename)
    shutil.copy2(baseline_path, wip_path)

    # 3. Data Ingestion & Stacking (Shadow Calculations)
    logging.info("Stacking %s Refund files for Module A...", len(type1_files))
    refunds_all = load_and_stack_files(type1_files, FILE_CONFIG['type_1']['header_keys'])
    logging.info("Stacking %s Consumption files for Module A...", len(type2_files))
    consumption_all = load_and_stack_files(type2_files, FILE_CONFIG['type_2']['header_keys'])

    account_statement_df = pd.concat([consumption_all, refunds_all], ignore_index=True)
    account_statement_df['Internal_Key_R'] = account_statement_df.iloc[:, 1].fillna('') + account_statement_df.iloc[:, 4].fillna('')

    logging.info("Stacking %s Channel Order files for Module B...", len(type3_files))
    channel_orders_raw = load_and_stack_files(type3_files, FILE_CONFIG['type_3']['header_keys'])
    channel_orders_filtered = channel_orders_raw[~channel_orders_raw.iloc[:, 9].isin(['预授权申请', '预授权撤销'])]

    # Module B Source 2: Re-validation
    special_orders_df = pd.read_excel(
        baseline_path,
        sheet_name='特殊的渠道订单',
        **READ_EXCEL_TEXT_KWARGS,
    )
    valid_from_special = pd.DataFrame()
    if not special_orders_df.empty:
        special_orders_df['Calc_AI_Key'] = special_orders_df.iloc[:, 3].fillna('') + special_orders_df.iloc[:, 4].fillna('')
        valid_from_special = special_orders_df[special_orders_df['Calc_AI_Key'].isin(account_statement_df['Internal_Key_R'])]

    # 4. Writing to Excel
    wb = load_workbook(wip_path)

    # Update Account Statement Sheet
    ws_acc = wb['账户流水']
    ws_acc.delete_rows(2, ws_acc.max_row) # Clear Old Data

    # Prepare workbook sheets used for writes and shadow lookups.
    keys_to_remove_from_statement = set()
    ws_chan = wb['渠道订单']
    ws_chan.delete_rows(2, ws_chan.max_row)

    ws_spec = wb['特殊的渠道订单']
    if ws_spec.max_row > 1:
        ws_spec.delete_rows(2, ws_spec.max_row)
    ws_payout = wb['打款币种']
    ws_a07 = wb['二级商户号映射表-A07']

    # Prepare the channel order payload before writing formulas to Excel.
    # Prepare Target Data Pool
    target_data = []
    for _, row in channel_orders_filtered.iterrows():
        new_row = list(row.iloc[0:4]) + list(row.iloc[5:34]) + [row.iloc[4]]
        target_data.append(new_row)
    if not valid_from_special.empty:
        for _, row in valid_from_special.iterrows():
            target_data.append(list(row.iloc[0:34]))

    # Load mapping sheets so dependent Excel formulas can be shadow-calculated in Python.
    mapping_sheets = {
        '打款币种': pd.read_excel(wip_path, sheet_name='打款币种', **READ_EXCEL_TEXT_KWARGS),
        '渠道名称': pd.read_excel(wip_path, sheet_name='渠道名称', **READ_EXCEL_TEXT_KWARGS),
        '二级商户号映射表-A01': pd.read_excel(wip_path, sheet_name='二级商户号映射表-A01', **READ_EXCEL_TEXT_KWARGS),
        '二级商户号映射表-A07': pd.read_excel(wip_path, sheet_name='二级商户号映射表-A07', **READ_EXCEL_TEXT_KWARGS),
    }

    # Build Python-side lookup maps used by the channel order resolution flow.
    payout_lookup = build_lookup_map(mapping_sheets['打款币种'], key_col_idx=6, value_col_idx=5)
    chan_map = build_lookup_map(mapping_sheets['渠道名称'], key_col_idx=0, value_col_idx=1)
    a01_lookup = build_lookup_map(mapping_sheets['二级商户号映射表-A01'], key_col_idx=0, value_col_idx=1)
    a07_lookup = build_lookup_map(mapping_sheets['二级商户号映射表-A07'], key_col_idx=0, value_col_idx=2)
    pending_payout_keys = set()
    pending_a07_keys = set()
    payout_rows_added = []
    a07_rows_added = []
    special_rows_added = []

    # Resolve channel-order exceptions and maintain missing payout-currency mappings.
    curr_row = 2
    exceptions_moved = 0

    for data_row in target_data:
        col_f, col_d, col_e = data_row[5], data_row[3], data_row[4]
        ai_key = f"{col_d}{col_e}"
        ap_val = str(chan_map.get(col_f, "")).strip()
        is_na = account_statement_df[account_statement_df['Internal_Key_R'] == ai_key].empty

        if is_na and ap_val.lower() not in ["paypal", "afterpay直连"]:
            # Move unmatched channel orders into the special-order sheet for follow-up.
            next_spec = ws_spec.max_row + 1
            for c_idx, val in enumerate(data_row, start=1):
                ws_spec.cell(row=next_spec, column=c_idx).value = to_excel_cell_value(val)
            ws_spec.cell(row=next_spec, column=35).value = f"=D{next_spec}&E{next_spec}"
            ws_spec.cell(row=next_spec, column=36).value = f"=XLOOKUP(A{next_spec},账户流水!$R:$R,账户流水!$R:$R)"
            special_rows_added.append([
                next_spec,
                get_data_row_value(data_row, 3),
                get_data_row_value(data_row, 1),
                ap_val,
                ai_key,
            ])
            keys_to_remove_from_statement.add(ai_key)
            exceptions_moved += 1
            continue

        # Shadow-calculate dependent formula values before writing the worksheet row.
        ah_value = get_data_row_value(data_row, 34)
        aj_value = get_data_row_value(data_row, 14)
        ab_value = get_data_row_value(data_row, 28)

        # Add missing A07 secondary-merchant mappings so AQ can be completed manually later.
        if ap_val == "A07" and ah_value and ah_value not in a07_lookup and ah_value not in pending_a07_keys:
            insert_row = append_a07_mapping_row(ws_a07, ah_value)
            a07_lookup[ah_value] = ""
            pending_a07_keys.add(ah_value)
            a07_rows_added.append([insert_row, ah_value])

        aq_value = resolve_aq_value(ap_val, ah_value, ab_value, a01_lookup, a07_lookup)
        ar_value = f"{ap_val}{aq_value}{aj_value}"

        # Add missing payout-currency mapping rows so later workbook lookups can resolve.
        if ap_val and aq_value and aj_value and ar_value not in payout_lookup and ar_value not in pending_payout_keys:
            insert_row = append_payout_currency_row(ws_payout, ap_val, ah_value, aj_value)
            payout_lookup[ar_value] = ""
            pending_payout_keys.add(ar_value)
            payout_rows_added.append([insert_row, ap_val, ah_value, aj_value, ar_value])

        # Write the source row first, then preserve the workbook formulas for auditing.
        # Write to Main Channel Orders
        for c_idx, val in enumerate(data_row, start=1):
            ws_chan.cell(row=curr_row, column=c_idx).value = to_excel_cell_value(val)

        # Write the live workbook formulas used for downstream human verification.
        ws_chan.cell(row=curr_row, column=35).value = f"=D{curr_row}&E{curr_row}"
        ws_chan.cell(row=curr_row, column=36).value = f"=N{curr_row}"
        ws_chan.cell(row=curr_row, column=37).value = f'=IF(I{curr_row}="退款", -M{curr_row}*(1-0.032), M{curr_row}*(1-0.032))'
        ws_chan.cell(row=curr_row, column=38).value = f"=XLOOKUP(AR{curr_row}, 打款币种!$G:$G, 打款币种!$F:$F)"
        ws_chan.cell(row=curr_row, column=39).value = f"=VLOOKUP(AI{curr_row}, 账户流水!$R:$T, 2, 0)"
        ws_chan.cell(row=curr_row, column=40).value = f"=VLOOKUP(AI{curr_row}, 账户流水!$R:$T, 3, 0)"
        ws_chan.cell(row=curr_row, column=41).value = f"=IF(AL{curr_row}=AM{curr_row}, \"是\", \"否\")"
        ws_chan.cell(row=curr_row, column=42).value = f"=VLOOKUP(F{curr_row}, 渠道名称!$A:$B, 2, 0)"

        aq_f = (f'IF(AP{curr_row}="2号通道", XLOOKUP(AH{curr_row}, \'二级商户号映射表-A01\'!$A:$A, \'二级商户号映射表-A01\'!$B:$B), '
                f'IF(AP{curr_row}="A07", XLOOKUP(AH{curr_row}, \'二级商户号映射表-A07\'!$A:$A, \'二级商户号映射表-A07\'!$C:$C) & AH{curr_row}, '
                f'IF(AP{curr_row}="7号通道", AB{curr_row}, "")))')
        ws_chan.cell(row=curr_row, column=43).value = f"={aq_f}"
        ws_chan.cell(row=curr_row, column=44).value = f"=AP{curr_row}&AQ{curr_row}&AJ{curr_row}"
        curr_row += 1

    # Remove statement rows that were pushed into the special-order sheet.
    # Apply Cleanup & Final Statement Write
    if keys_to_remove_from_statement:
        account_statement_df = account_statement_df[~account_statement_df['Internal_Key_R'].isin(keys_to_remove_from_statement)]

    # Rewrite the account statement with formulas only where there is actual source data.
    for r_idx, row in enumerate(account_statement_df.values, start=2):
        # Only write the row and formula if the first few columns aren't empty
        if pd.notna(row[0]) and str(row[0]).strip() != "":
            for c_idx, value in enumerate(row[:-1], start=1):
                ws_acc.cell(row=r_idx, column=c_idx).value = to_excel_cell_value(value)

            # Only apply formula if there is data to calculate
            ws_acc.cell(row=r_idx, column=13).value = f"=K{r_idx}-L{r_idx}"
            ws_acc.cell(row=r_idx, column=18).value = f"=B{r_idx}&E{r_idx}"
            ws_acc.cell(row=r_idx, column=19).value = f"=G{r_idx}"
            ws_acc.cell(row=r_idx, column=20).value = f"=M{r_idx}"

    logging.info(f"Processing finished. Exceptions: {exceptions_moved}")
    logging.info(
        "\n%s",
        format_summary_table(
            "Added rows in 二级商户号映射表-A07",
            [
                "Row",
                f"A-{get_sheet_header(ws_a07, 1)}",
            ],
            a07_rows_added,
        ),
    )
    logging.info(
        "\n%s",
        format_summary_table(
            "Added rows in 打款币种",
            [
                "Row",
                f"B-{get_sheet_header(ws_payout, 2)}",
                f"D-{get_sheet_header(ws_payout, 4)}",
                f"E-{get_sheet_header(ws_payout, 5)}",
                f"G-{get_sheet_header(ws_payout, 7)}",
            ],
            payout_rows_added,
        ),
    )
    logging.info(
        "\n%s",
        format_summary_table(
            "Added rows in 特殊的渠道订单",
            ["Row", "C-交易流水号", "A-渠道订单号", "AP-通道名称", "AI-Key"],
            special_rows_added,
        ),
    )
    wb.save(wip_path)
    final_path = wip_path.replace("-wip.xlsx", ".xlsx")
    if os.path.exists(final_path): os.remove(final_path)
    os.rename(wip_path, final_path)
    logging.info(f"COMPLETED. File: {final_path}")


if __name__ == "__main__":
    main()
